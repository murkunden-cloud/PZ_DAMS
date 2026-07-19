import os
import json
import tempfile
import openpyxl
from openpyxl.styles import PatternFill

def get_audit_log_path(app_dir):
    return os.path.join(app_dir, "runtime", "web_edits_log.json")

def log_edit(app_dir, sheet_name, row_number, col_index):
    """Log an edit to the audit log."""
    log_path = get_audit_log_path(app_dir)
    log_data = {}
    if os.path.exists(log_path):
        try:
            with open(log_path, 'r') as f:
                log_data = json.load(f)
        except Exception:
            pass
            
    if sheet_name not in log_data:
        log_data[sheet_name] = {}
        
    row_str = str(row_number)
    if row_str not in log_data[sheet_name]:
        log_data[sheet_name][row_str] = []
        
    if col_index not in log_data[sheet_name][row_str]:
        log_data[sheet_name][row_str].append(col_index)
        
    with open(log_path, 'w') as f:
        json.dump(log_data, f)

def log_row_added(app_dir, sheet_name, row_number, max_cols=30):
    """Log an entire row as added/edited."""
    for c in range(1, max_cols + 1):
        log_edit(app_dir, sheet_name, row_number, c)

def clear_audit_log(app_dir):
    """Clear the audit log (e.g. for a new month)."""
    log_path = get_audit_log_path(app_dir)
    if os.path.exists(log_path):
        os.remove(log_path)

def generate_highlighted_export(master_file, app_dir, output_prefix="Updated_DC"):
    """
    Reads the audit log and highlights exactly those cells in the master file.
    Returns the path to the temporary highlighted Excel file.
    """
    if not os.path.exists(master_file):
        raise FileNotFoundError("Master file not found.")

    wb = openpyxl.load_workbook(master_file)
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    
    log_path = get_audit_log_path(app_dir)
    log_data = {}
    if os.path.exists(log_path):
        try:
            with open(log_path, 'r') as f:
                log_data = json.load(f)
        except Exception:
            pass

    for sheet_name, rows in log_data.items():
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_str, cols in rows.items():
                try:
                    r = int(row_str)
                    for c in cols:
                        sheet.cell(row=r, column=c).fill = yellow_fill
                except ValueError:
                    continue

    # Save to a temporary file
    temp_dir = tempfile.gettempdir()
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_path = os.path.join(temp_dir, f"{output_prefix}_{timestamp}.xlsx")
    
    wb.save(temp_path)
    wb.close()
    
    return temp_path
