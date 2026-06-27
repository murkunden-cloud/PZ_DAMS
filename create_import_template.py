import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_import_template():
    """Create Excel template for monthly case import"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet structures based on META
    sheets_config = {
        "4DC": {
            "title": "Minor DC - Statewise (Class I/II)",
            "headers": [
                "Sr No", "Zone", "Circle", "Division", "CPF No", "Employee Name",
                "Designation", "Pay Group", "Working Place", "DC No",
                "Dispatch Reference", "Last Updated", "Remarks"
            ],
            "data_start": 5
        },
        "5DC": {
            "title": "Minor DC - Circlewise (Class III/IV)",
            "headers": [
                "Sr No", "Circle", "Division", "CPF No", "Employee Name",
                "Designation", "Pay Group", "Working Place", "DC No",
                "Dispatch Reference", "Last Updated", "Remarks"
            ],
            "data_start": 5
        },
        "20DC": {
            "title": "Major DC Chargesheet - Statewise",
            "headers": [
                "Sr No", "Zone", "Circle", "Division", "CPF No", "Employee Name",
                "Designation", "Pay Group", "Working Place", "DC No",
                "Dispatch Reference", "Last Updated", "Remarks"
            ],
            "data_start": 6
        },
        "21DC": {
            "title": "Major DC Chargesheet - Circlewise",
            "headers": [
                "Sr No", "Circle", "Division", "CPF No", "Employee Name",
                "Designation", "Pay Group", "Working Place", "DC No",
                "Dispatch Reference", "Last Updated", "Remarks"
            ],
            "data_start": 5
        },
        "6DC": {
            "title": "Minor DC - Consolidated (4DC+5DC)",
            "headers": [
                "Sr No", "Circle", "Division", "CPF No", "Employee Name",
                "Designation", "Pay Group", "Working Place", "DC No",
                "Dispatch Reference", "DC Record No", "DC Record Date", "Remarks"
            ],
            "data_start": 4
        },
        "22DC": {
            "title": "Major DC All Cases - Statewise",
            "headers": [
                "Sr No", "Zone", "Circle", "Division", "CPF No", "Employee Name",
                "Designation", "Pay Group", "Working Place", "DC No",
                "Charge Sheet No & Date", "DC Record No", "DC Record Date", "Remarks"
            ],
            "data_start": 6
        },
        "23DC": {
            "title": "Major DC All Cases - Circlewise",
            "headers": [
                "Sr No", "Circle", "Division", "CPF No", "Employee Name",
                "Designation", "Pay Group", "Working Place", "DC No",
                "Charge Sheet No & Date", "DC Record No", "DC Record Date", "Remarks"
            ],
            "data_start": 5
        }
    }
    
    # Create sheets
    for sheet_name, config in sheets_config.items():
        # Remove default sheet if it exists
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Add title
        ws.merge_cells('A1:M1')
        ws['A1'] = config["title"]
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='2C3E50')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add instruction
        ws.merge_cells('A2:M2')
        ws['A2'] = f"Monthly Import Template - {datetime.now().strftime('%B %Y')}"
        ws['A2'].font = Font(name='Arial', size=10, italic=True, color='7F8C8D')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        header_row = config["data_start"] - 1
        for col_idx, header in enumerate(config["headers"], start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths
        column_widths = {
            1: 8,   # Sr No
            2: 12,  # Zone/Circle
            3: 12,  # Circle/Division
            4: 12,  # Division
            5: 12,  # CPF No
            6: 25,  # Employee Name
            7: 20,  # Designation
            8: 12,  # Pay Group
            9: 20,  # Working Place
            10: 15, # DC No
            11: 20, # Dispatch Reference / Charge Sheet No & Date
            12: 15, # DC Record No / Last Updated
            13: 15, # DC Record Date / Remarks
            14: 30  # Remarks (for sheets with DC Record columns)
        }
        
        for col_idx, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # Set row heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[header_row].height = 22
        ws.row_dimensions[header_row + 1].height = 20
    
    # Save template
    template_path = "Monthly_Case_Import_Template_v2.xlsx"
    wb.save(template_path)
    print(f"Template created: {template_path}")
    print("\nTemplate includes sheets: 4DC, 5DC, 20DC, 21DC, 6DC, 22DC, 23DC")
    print("Fill in data starting from the header row and save as Excel file for import.")
    print("\nNew columns added:")
    print("- 6DC: DC Record No (col 11), DC Record Date (col 12)")
    print("- 22DC, 23DC: DC Record No (col 11), DC Record Date (col 12)")

if __name__ == "__main__":
    create_import_template()
