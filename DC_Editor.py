import json
import os
import shutil
import re
from datetime import datetime

try:
    import openpyxl
    import pandas as pd
except ImportError as exc:
    raise RuntimeError("openpyxl and pandas are required. Install them with: pip install openpyxl flask pandas") from exc

from DC_BackupManager import DCBackupManager
from DC_DataLoader import (
    CLOSED_DB,
    EXTRACTED_DC_HEADER,
    normalize_cpf,
    extract_dc,
    has_cpf_value,
    row_to_strings,
    safe_text,
)

# Try to import database manager
try:
    from DC_DatabaseManager import db_manager
    DATABASE_AVAILABLE = True
except ImportError:
    DATABASE_AVAILABLE = False


def normalize_header_key(text):
    text = str(text).lower().strip()
    if "cpf" in text:
        return "cpfno"
    if "name" in text:
        return "name"
    if "designation" in text or "desg" in text:
        return "designation"
    if "pay group" in text or "pay  group" in text or "pay gr" in text:
        return "paygroup"
    if "retirement" in text or "birth" in text or "retire" in text or "dob" in text:
        return "retirement"
    if re.search(r'\b(?:place|hq|working)\b', text):
        return "place"
    if "theft" in text:
        return "theft"
        
    if "suspension order" in text:
        return "susp_order"
    if "effective date" in text:
        return "susp_date"
        return "theft"
    
    # Specific case state keys first to avoid collisions
    if "without" in text:
        return "close_ref"
    if "punishment" in text or "penalty" in text:
        return "punishment"
    if "chargesheet" in text or "charge sheet" in text:
        return "chargesheet"
        
    if "dispatch" in text or "letter" in text:
        return "dispatch"
    if "vs" in text or "cio" in text or "reference" in text or "ref" in text:
        return "vs_ref"
    if "facts" in text or "brief" in text or "charges" in text:
        return "facts"
    if "outcome" in text:
        return "outcome"
    if "status" in text or "explanation" in text:
        return "status"
    if "remarks" in text or "note" in text:
        return "remarks"
    return text


def split_dispatch_date(text):
    text = str(text).strip()
    # matches dtd/dt/dated/dt. followed by optional separator (- or :) and then dd.mm.yyyy or dd/mm/yyyy or dd-mm-yyyy
    pattern = re.compile(r'\b(?:dtd|dt\.?|dated|dt|dtd\.)\s*[-:]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})', re.IGNORECASE)
    match = pattern.search(text)
    if match:
        date_val = match.group(1)
        num_val = pattern.sub('', text).strip()
        num_val = re.sub(r'\s*[dD]td\s*$', '', num_val).strip()
        num_val = num_val.rstrip(", /.-").strip()
        return num_val, date_val
        
    fallback_pattern = re.compile(r'\b(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\s*$', re.IGNORECASE)
    match = fallback_pattern.search(text)
    if match:
        date_val = match.group(1)
        num_val = fallback_pattern.sub('', text).strip()
        num_val = num_val.rstrip(", /.-").strip()
        return num_val, date_val
        
    return text, ""



class DCEditor:
    def __init__(self, loader):
        self.loader = loader
        self.backups = DCBackupManager(loader.dc_file)

    def save_case_to_database(self, case_data):
        """Save case data to database"""
        if not DATABASE_AVAILABLE:
            return False
        
        try:
            case_id = case_data.get('case_id')
            # If case_id exists, it's an update (but we need the numeric id)
            # For now, we'll just insert new cases
            if not case_id:
                # Generate a case ID
                from datetime import datetime
                sheet = case_data.get('sheet_origin', 'UNKNOWN')
                cpf = case_data.get('cpf_no', 'UNKNOWN')
                case_data['case_id'] = f"{sheet}_{cpf}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            return db_manager.insert_case(case_data)
        except Exception as e:
            print(f"Error saving case to database: {e}")
            return False

    def open_wb(self, data_only=False, filepath=None):
        target = filepath if filepath else self.loader.dc_file
        return openpyxl.load_workbook(target, data_only=data_only)

    def save_wb(self, workbook, modified_sheets=None):
        workbook.save(self.loader.dc_file)
        if modified_sheets:
            for sheet in modified_sheets:
                self.loader.clear_sheet_cache(sheet)
        else:
            self.loader.clear_cache()
            
        try:
            import threading
            def sync_db():
                try:
                    from migrate_cases import migrate_cases
                    migrate_cases()
                    self.loader.clear_cache()
                except Exception as e:
                    print(f"Background DB sync failed: {e}")
            threading.Thread(target=sync_db, daemon=True).start()
        except Exception:
            pass

    def archive_closed_cases(self, auto_delete=True):
        self.backups.backup(reason="archive_closed_cases_pre")
        archive_file = os.path.join(os.path.dirname(self.loader.dc_file), "closed_cases_master.xlsx")
        
        is_first_run = False
        if not os.path.exists(archive_file):
            shutil.copy2(self.loader.dc_file, archive_file)
            is_first_run = True

        closed_sheets = []
        for s, m in self.loader.meta.items():
            t = m.get("type", "")
            if "close" in t or "revoke" in t:
                closed_sheets.append(s)

        source_wb = self.open_wb()
        archive_wb = openpyxl.load_workbook(archive_file)
        
        try:
            for s in closed_sheets:
                actual_src = self.get_actual_sheet_name(s)
                if not actual_src or actual_src not in source_wb.sheetnames:
                    continue
                    
                meta = self.loader.get_sheet_meta(s)
                data_start = meta.get("data_start", 5)
                
                src_sheet = source_wb[actual_src]
                arc_sheet = archive_wb[actual_src] if actual_src in archive_wb.sheetnames else None
                
                if not arc_sheet:
                    continue
                    
                max_src_row = src_sheet.max_row
                
                if not is_first_run and max_src_row >= data_start:
                    # Append rows to archive
                    arc_max = arc_sheet.max_row
                    for r_idx in range(data_start, max_src_row + 1):
                        arc_max += 1
                        for c_idx in range(1, src_sheet.max_column + 1):
                            val = src_sheet.cell(row=r_idx, column=c_idx).value
                            arc_sheet.cell(row=arc_max, column=c_idx, value=val)
                        # Fix Sr. No (usually col 1)
                        arc_sheet.cell(row=arc_max, column=1, value=arc_max - data_start + 1)
                
                if auto_delete and max_src_row >= data_start:
                    # Delete rows from working file from bottom up to avoid shifting index issues
                    for r_idx in range(max_src_row, data_start - 1, -1):
                        src_sheet.delete_rows(r_idx)
                        
            archive_wb.save(archive_file)
            if auto_delete:
                source_wb.save(self.loader.dc_file)
                self.loader.clear_cache()
                
            return archive_file
            
        finally:
            source_wb.close()
            archive_wb.close()

    def get_extract_column(self, worksheet):
        for row in range(1, min(worksheet.max_row, 5) + 1):
            for col in range(1, worksheet.max_column + 1):
                if str(worksheet.cell(row, col).value).strip() == EXTRACTED_DC_HEADER:
                    return col
        new_col = worksheet.max_column + 1
        worksheet.cell(1, new_col, EXTRACTED_DC_HEADER)
        return new_col

    def get_actual_sheet_name(self, sheet_name):
        return self.loader.meta.get(sheet_name, {}).get("actual_name", sheet_name)

    def view_sheet(self, sheet_name, filepath=None):
        workbook = self.open_wb(filepath=filepath)
        actual_sheet = self.get_actual_sheet_name(sheet_name)
        try:
            if actual_sheet not in workbook.sheetnames:
                raise KeyError(f"Sheet '{sheet_name}' not found")
            worksheet = workbook[actual_sheet]
            meta = self.loader.get_sheet_meta(sheet_name)
            headers = self.get_column_labels(sheet_name, workbook=workbook)
            return {
                "sheet": sheet_name,
                "title": meta.get("title", ""),
                "rows": worksheet.max_row,
                "cols": worksheet.max_column,
                "cpf_col": meta.get("cpf_col"),
                "dc_col": meta.get("dc_col"),
                "data_start": meta.get("data_start"),
                "headers": headers,
            }
        finally:
            workbook.close()

    def get_column_labels(self, sheet_name, workbook=None):
        should_close = workbook is None
        workbook = workbook or self.open_wb()
        actual_sheet = self.get_actual_sheet_name(sheet_name)
        try:
            if actual_sheet not in workbook.sheetnames:
                return []
            worksheet = workbook[actual_sheet]
            meta = self.loader.get_sheet_meta(sheet_name)
            header_scan_rows = meta.get("data_start", 4) - 1
            labels = []
            for col in range(1, worksheet.max_column + 1):
                label = ""
                for row in range(1, header_scan_rows + 1):
                    value = worksheet.cell(row, col).value
                    if value not in (None, ""):
                        text = str(value).strip()
                        if text and text not in (str(col),):
                            label = text
                if label:
                    labels.append({"column": col, "header": label})
            
            # Smart fallback: if mostly empty (happens when data_start is unconfigured on 35DC sheets)
            if sum(1 for l in labels if l['header']) < 3:
                best_row = 1
                max_cols = 0
                for r in range(1, min(10, worksheet.max_row + 1)):
                    cols_with_data = sum(1 for c in range(1, worksheet.max_column + 1) if worksheet.cell(r, c).value not in (None, ""))
                    if cols_with_data > max_cols:
                        max_cols = cols_with_data
                        best_row = r
                
                labels = []
                for col in range(1, worksheet.max_column + 1):
                    label = ""
                    for row in range(1, best_row + 1):
                        value = worksheet.cell(row, col).value
                        if value not in (None, ""):
                            label = str(value)
                    labels.append({"column": col, "header": label.strip()})
            
            if not labels:
                # fallback to first-row values when header detection fails
                for col in range(1, worksheet.max_column + 1):
                    value = worksheet.cell(1, col).value
                    label = "" if value is None else str(value).strip()
                    if label:
                        labels.append({"column": col, "header": label})
            return labels
        finally:
            if should_close:
                workbook.close()

    def get_sheet_field_map(self, sheet_name):
        labels = self.get_column_labels(sheet_name)
        field_map = {}
        for item in labels:
            col = item["column"]
            header = item["header"]
            key = normalize_header_key(header)
            if key not in field_map:
                field_map[key] = col
        return field_map

    def get_employee_by_cpf(self, cpf):
        emp_df = self.loader.load_emp()
        target = normalize_cpf(cpf)
        if not emp_df.empty and "CPFNO_NORM" in emp_df.columns:
            match = emp_df[emp_df["CPFNO_NORM"] == target]
            if not match.empty:
                return match.iloc[0].to_dict()
        return None

    def get_record(self, sheet_name, row_number, workbook=None, filepath=None):
        labels = {item["column"]: item["header"] for item in self.get_column_labels(sheet_name)}
        frame = self.loader.load_dc_sheet(sheet_name, use_cache=True)
        if frame is not None and 1 <= int(row_number) <= len(frame):
            row = frame.iloc[int(row_number) - 1]
            cells = []
            record_map = {}
            columns_data = []
            
            field_map = self.get_sheet_field_map(sheet_name)
            disp_col = field_map.get("dispatch")
            cpf_col = field_map.get("cpfno")
            
            # Get CPF and fetch employee details from master
            cpf_val = ""
            emp_details = {}
            if cpf_col and cpf_col <= len(row):
                cpf_val = normalize_cpf(row.iloc[cpf_col - 1])
                if cpf_val:
                    emp_details = self.get_employee_by_cpf(cpf_val) or {}
            
            for col in range(1, frame.shape[1] + 1):
                value = row.iloc[col - 1] if col - 1 < len(row) else ""
                val_text = safe_text(value)
                
                is_disp = (col == disp_col)
                disp_num, disp_date = ("", "")
                if is_disp:
                    disp_num, disp_date = split_dispatch_date(val_text)
                    
                header_name = labels.get(col, f"Column {col}")
                columns_data.append(
                    {
                        "column": col,
                        "header": header_name,
                        "value": val_text,
                        "is_dispatch": is_disp,
                        "disp_num": disp_num,
                        "disp_date": disp_date,
                        "is_extracted": (header_name == EXTRACTED_DC_HEADER),
                    }
                )
                if pd.isna(value):
                    continue
                text = safe_text(value)
                if not text or text.startswith("="):
                    continue
                cells.append(
                    {
                        "column": col,
                        "label": labels.get(col, f"Column {col}"),
                        "value": text,
                    }
                )
                record_map[col] = text
            row_values = row_to_strings(list(row.iloc[: min(frame.shape[1], 20)]))
            return {
                "sheet": sheet_name,
                "row_number": int(row_number),
                "cells": cells,
                "record_map": record_map,
                "row_values": row_values,
                "columns_data": columns_data,
                "edit_text": "\n".join(f"{col}={value}" for col, value in record_map.items()),
                "employee_details": emp_details,
            }

        should_close = workbook is None
        workbook = workbook or self.open_wb(data_only=True, filepath=filepath)
        actual_sheet = self.get_actual_sheet_name(sheet_name)
        try:
            if actual_sheet not in workbook.sheetnames:
                raise KeyError(f"Sheet '{sheet_name}' not found")
            worksheet = workbook[actual_sheet]
            cells = []
            record_map = {}
            columns_data = []
            
            field_map = self.get_sheet_field_map(sheet_name)
            disp_col = field_map.get("dispatch")
            cpf_col = field_map.get("cpfno")
            
            # Get CPF and fetch employee details from master
            cpf_val = ""
            emp_details = {}
            if cpf_col:
                cpf_cell_value = worksheet.cell(int(row_number), cpf_col).value
                cpf_val = normalize_cpf(cpf_cell_value)
                if cpf_val:
                    emp_details = self.get_employee_by_cpf(cpf_val) or {}
            
            for col in range(1, worksheet.max_column + 1):
                value = worksheet.cell(int(row_number), col).value
                text = "" if value is None else str(value).strip()
                
                is_disp = (col == disp_col)
                disp_num, disp_date = ("", "")
                if is_disp:
                    disp_num, disp_date = split_dispatch_date(text)
                    
                header_name = labels.get(col, f"Column {col}")
                columns_data.append(
                    {
                        "column": col,
                        "header": header_name,
                        "value": text,
                        "is_dispatch": is_disp,
                        "disp_num": disp_num,
                        "disp_date": disp_date,
                        "is_extracted": (header_name == EXTRACTED_DC_HEADER),
                    }
                )
                if not text or text.startswith("="):
                    continue
                cells.append(
                    {
                        "column": col,
                        "label": labels.get(col, f"Column {col}"),
                        "value": text,
                    }
                )
                record_map[col] = text
            return {
                "sheet": sheet_name,
                "row_number": int(row_number),
                "cells": cells,
                "record_map": record_map,
                "columns_data": columns_data,
                "edit_text": "\n".join(f"{col}={value}" for col, value in record_map.items()),
                "employee_details": emp_details,
            }
        finally:
            if should_close:
                workbook.close()

    def find_by_cpf(self, sheet_name, cpfno, workbook=None, filepath=None):
        should_close = workbook is None
        workbook = workbook or self.open_wb(filepath=filepath)
        actual_sheet = self.get_actual_sheet_name(sheet_name)
        try:
            if actual_sheet not in workbook.sheetnames:
                return []
            worksheet = workbook[actual_sheet]
            meta = self.loader.get_sheet_meta(sheet_name)
            cpf_col = meta.get("cpf_col", 3)
            start_row = meta.get("data_start", 4)
            target = normalize_cpf(cpfno)
            results = []
            for row_number in range(start_row, worksheet.max_row + 1):
                value = worksheet.cell(row_number, cpf_col).value
                if target and normalize_cpf(value) == target:
                    row_values = [worksheet.cell(row_number, col).value for col in range(1, worksheet.max_column + 1)]
                    results.append(
                        {
                            "row_number": row_number,
                            "sheet": sheet_name,
                            "row_data": row_to_strings(row_values),
                        }
                    )
            return results
        finally:
            if should_close:
                workbook.close()

    def add_record(self, sheet_name, record, reason="add_record"):
        self.backups.backup(reason=reason)
        
        # Auto-fill employee details if CPF is present
        field_map = self.get_sheet_field_map(sheet_name)
        cpf_col = field_map.get("cpfno")
        if cpf_col:
            cpf_val = record.get(cpf_col) or record.get(str(cpf_col))
            if cpf_val:
                emp = self.get_employee_by_cpf(cpf_val)
                if emp:
                    # Map employee fields
                    name_col = field_map.get("name")
                    desg_col = field_map.get("designation")
                    if name_col and not record.get(name_col) and not record.get(str(name_col)):
                        if desg_col:
                            record[name_col] = emp.get("EmployeeName", "")
                        else:
                            record[name_col] = f"{emp.get('EmployeeName', '')} {emp.get('Designation', '')}"
                    if desg_col and not record.get(desg_col) and not record.get(str(desg_col)):
                        record[desg_col] = emp.get("Designation", "")
                    
                    pg_col = field_map.get("paygroup")
                    if pg_col and not record.get(pg_col) and not record.get(str(pg_col)):
                        desg = emp.get("Designation", "")
                        pg = self.loader.get_designation_paygrp(desg)
                        if not pg:
                            pg = "IV"
                        record[pg_col] = pg
                        
                    place_col = field_map.get("place")
                    if place_col and not record.get(place_col) and not record.get(str(place_col)):
                        record[place_col] = emp.get("PresentOffice", "")

        workbook = self.open_wb()
        actual_sheet = self.get_actual_sheet_name(sheet_name)
        try:
            if actual_sheet not in workbook.sheetnames:
                raise KeyError(f"Sheet '{sheet_name}' not found")
            worksheet = workbook[actual_sheet]
            meta = self.loader.get_sheet_meta(sheet_name)
            start_row = meta.get("data_start", 4)
            next_row = start_row
            for row_number in range(start_row, worksheet.max_row + 2):
                empty = all(not worksheet.cell(row_number, col).value for col in range(1, 6))
                if empty:
                    next_row = row_number
                    break
            
            # Auto-calculate next Sr. No. if column 1 exists and is not explicitly set
            if next_row > start_row:
                try:
                    prev_val = worksheet.cell(next_row - 1, 1).value
                    if prev_val and str(prev_val).strip().isdigit():
                        worksheet.cell(next_row, 1, int(str(prev_val).strip()) + 1)
                except:
                    pass
            else:
                worksheet.cell(next_row, 1, 1)

            from export_utils import log_edit
            from DC_DataLoader import APP_DIR
            for col, value in record.items():
                if int(col) == 1:
                    continue  # skip manual Sr. No overwrite
                worksheet.cell(next_row, int(col), value)
                log_edit(APP_DIR, sheet_name, next_row, int(col))
            
            dc_col = meta.get("dc_col")
            if dc_col and record.get(dc_col):
                worksheet.cell(next_row, self.get_extract_column(worksheet), extract_dc(record[dc_col]))
            
            self.save_wb(workbook, modified_sheets=[sheet_name])
            return {"sheet": sheet_name, "row_number": next_row}
        finally:
            workbook.close()

    def get_sheet_preview(self, sheet_name, max_rows=50, filepath=None):
        frame = self.loader.load_dc_sheet(sheet_name, use_cache=True)
        if frame is None:
            return []
        meta = self.loader.get_sheet_meta(sheet_name)
        start_row = max(meta.get("data_start", 4) - 1, 0)
        preview = []
        for idx, row in frame.iloc[start_row:start_row + max_rows].iterrows():
            preview.append({
                "row_number": int(idx) + 1,
                "values": row_to_strings(list(row)),
            })
        return preview

    def update_record(self, sheet_name, row_number, record, reason="update_record"):
        self.backups.backup(reason=reason)
        workbook = self.open_wb()
        actual_sheet = self.get_actual_sheet_name(sheet_name)
        try:
            if actual_sheet not in workbook.sheetnames:
                raise KeyError(f"Sheet '{sheet_name}' not found")
            worksheet = workbook[actual_sheet]
            meta = self.loader.get_sheet_meta(sheet_name)
            from export_utils import log_edit
            from DC_DataLoader import APP_DIR
            for col, value in record.items():
                worksheet.cell(int(row_number), int(col), value)
                log_edit(APP_DIR, sheet_name, int(row_number), int(col))
            dc_col = meta.get("dc_col")
            if dc_col and record.get(dc_col):
                worksheet.cell(int(row_number), self.get_extract_column(worksheet), extract_dc(record[dc_col]))
            self.save_wb(workbook, modified_sheets=[sheet_name])
            return {"sheet": sheet_name, "row_number": int(row_number)}
        finally:
            workbook.close()

    def delete_record(self, sheet_name, row_number, reason="delete_record"):
        self.backups.backup(reason=reason)
        workbook = self.open_wb()
        actual_sheet = self.get_actual_sheet_name(sheet_name)
        try:
            if actual_sheet not in workbook.sheetnames:
                raise KeyError(f"Sheet '{sheet_name}' not found")
            worksheet = workbook[actual_sheet]
            worksheet.delete_rows(int(row_number))
            self.save_wb(workbook, modified_sheets=[sheet_name])
            return {"sheet": sheet_name, "row_number": int(row_number)}
        finally:
            workbook.close()

    def copy_record_to_sheet(self, source_sheet, row_number, target_sheet):
        """Duplicates a record from source_sheet to target_sheet mapping columns automatically."""
        self.backups.backup(reason=f"copy_case_{source_sheet}_to_{target_sheet}")
        workbook = self.open_wb()
        try:
            actual_src = self.get_actual_sheet_name(source_sheet)
            actual_tgt = self.get_actual_sheet_name(target_sheet)
            if actual_src not in workbook.sheetnames or actual_tgt not in workbook.sheetnames:
                return None
                
            src_ws = workbook[actual_src]
            tgt_ws = workbook[actual_tgt]
            
            src_meta = self.loader.get_sheet_meta(source_sheet)
            tgt_meta = self.loader.get_sheet_meta(target_sheet)
            
            src_fields = self.get_sheet_field_map(source_sheet)
            tgt_fields = self.get_sheet_field_map(target_sheet)
            
            row_idx = int(row_number)
            row_vals = [src_ws.cell(row_idx, col).value for col in range(1, src_ws.max_column + 1)]
            
            tgt_start = tgt_meta.get("data_start", 4)
            next_row = tgt_start
            for candidate in range(tgt_start, tgt_ws.max_row + 2):
                empty = all(not tgt_ws.cell(candidate, col).value for col in range(1, 6))
                if empty:
                    next_row = candidate
                    break
                    
            # Auto-calculate target Sr. No.
            prev_sr = 0
            if next_row > tgt_start:
                try:
                    prev_sr_val = tgt_ws.cell(next_row - 1, 1).value
                    if prev_sr_val and str(prev_sr_val).strip().isdigit():
                        prev_sr = int(str(prev_sr_val).strip())
                except:
                    pass
            tgt_ws.cell(next_row, 1, prev_sr + 1)
            
            # Map fields
            vals_map = {}
            for key, src_col in src_fields.items():
                if src_col <= len(row_vals):
                    vals_map[key] = row_vals[src_col - 1]
                    
            from export_utils import log_row_added
            from DC_DataLoader import APP_DIR
            for key, tgt_col in tgt_fields.items():
                val = vals_map.get(key)
                if val is not None:
                    tgt_ws.cell(next_row, tgt_col, val)
                    
            log_row_added(APP_DIR, target_sheet, next_row, max_cols=tgt_ws.max_column)
            self.save_wb(workbook, modified_sheets=[target_sheet])
            return next_row
        finally:
            workbook.close()


    def extract_all_dc(self):
        self.backups.backup(reason="extract_all_dc")
        workbook = self.open_wb()
        stats = []
        try:
            for sheet_name in self.loader.op_sheets:
                actual_sheet = self.get_actual_sheet_name(sheet_name)
                if actual_sheet not in workbook.sheetnames:
                    continue
                worksheet = workbook[actual_sheet]
                meta = self.loader.get_sheet_meta(sheet_name)
                cpf_col = meta.get("cpf_col", 3)
                dc_col = meta.get("dc_col", 10)
                start_row = meta.get("data_start", 4)
                extract_col = self.get_extract_column(worksheet)
                count = 0
                for row_number in range(start_row, worksheet.max_row + 1):
                    cpf = worksheet.cell(row_number, cpf_col).value
                    if has_cpf_value(cpf):
                        dispatch = worksheet.cell(row_number, dc_col).value
                        worksheet.cell(row_number, extract_col, extract_dc(dispatch))
                        count += 1
                stats.append({"sheet": sheet_name, "count": count})
            self.save_wb(workbook)
            return stats
        finally:
            workbook.close()

    def move_or_close_case(self, source_sheet, row_number, target_sheet, closure_details=None):
        self.backups.backup(reason=f"move_case_{source_sheet}_to_{target_sheet}")
        workbook = self.open_wb()
        try:
            actual_src = self.get_actual_sheet_name(source_sheet)
            actual_tgt = self.get_actual_sheet_name(target_sheet)
            if actual_src not in workbook.sheetnames:
                raise KeyError(f"Source sheet '{source_sheet}' not found")
            if actual_tgt not in workbook.sheetnames:
                raise KeyError(f"Target sheet '{target_sheet}' not found")
                
            src_ws = workbook[actual_src]
            tgt_ws = workbook[actual_tgt]
            
            src_meta = self.loader.get_sheet_meta(source_sheet)
            tgt_meta = self.loader.get_sheet_meta(target_sheet)
            
            src_fields = self.get_sheet_field_map(source_sheet)
            tgt_fields = self.get_sheet_field_map(target_sheet)
            
            row_idx = int(row_number)
            row_vals = [src_ws.cell(row_idx, col).value for col in range(1, src_ws.max_column + 1)]
            
            cpf_col = src_fields.get("cpfno")
            cpf_val = ""
            if cpf_col and cpf_col <= len(row_vals):
                cpf_val = normalize_cpf(row_vals[cpf_col - 1])
                
            emp = self.get_employee_by_cpf(cpf_val) if cpf_val else None
            
            tgt_start = tgt_meta.get("data_start", 4)
            next_row = tgt_start
            for candidate in range(tgt_start, tgt_ws.max_row + 2):
                empty = all(not tgt_ws.cell(candidate, col).value for col in range(1, 6))
                if empty:
                    next_row = candidate
                    break
                    
            # Auto-calculate target Sr. No.
            prev_sr = 0
            if next_row > tgt_start:
                try:
                    prev_sr_val = tgt_ws.cell(next_row - 1, 1).value
                    if prev_sr_val and str(prev_sr_val).strip().isdigit():
                        prev_sr = int(str(prev_sr_val).strip())
                except:
                    pass
            tgt_ws.cell(next_row, 1, prev_sr + 1)
            
            # Map fields
            vals_map = {}
            for key, src_col in src_fields.items():
                if src_col <= len(row_vals):
                    vals_map[key] = row_vals[src_col - 1]
                    
            if emp:
                vals_map["cpfno"] = emp.get("CPFNO", cpf_val)
                if "name" in tgt_fields and "designation" not in tgt_fields:
                    vals_map["name"] = f"{emp.get('EmployeeName', '')} {emp.get('Designation', '')}"
                else:
                    vals_map["name"] = emp.get("EmployeeName", "")
                    vals_map["designation"] = emp.get("Designation", "")
                
                # Guess pay group
                desg = emp.get("Designation", "")
                desg_lower = desg.lower()
                if any(k in desg_lower for k in ["chief", "superintending", "executive engineer", "ee", "additional"]):
                    pg = "I"
                elif any(k in desg_lower for k in ["assistant engineer", "dy executive", "dy.ee", "ae", "adee"]):
                    pg = "II"
                elif any(k in desg_lower for k in ["junior engineer", "je", "sub engineer", "operator", "technician", "clerk"]):
                    pg = "III"
                else:
                    pg = "IV"
                vals_map["paygroup"] = pg
                vals_map["place"] = emp.get("PresentOffice", "")
                
            if closure_details:
                for k, v in closure_details.items():
                    vals_map[k] = v
                    
            for key, tgt_col in tgt_fields.items():
                if tgt_col == 1:
                    continue
                val = vals_map.get(key)
                if val is not None:
                    tgt_ws.cell(next_row, tgt_col, val)
                    
            dc_col = tgt_meta.get("dc_col")
            if dc_col and dc_col in tgt_fields:
                dc_val = vals_map.get(dc_col)
                if dc_val:
                    tgt_ws.cell(next_row, self.get_extract_column(tgt_ws), extract_dc(dc_val))
                    
            # Delete row from source sheet
            src_ws.delete_rows(row_idx)
            
            self.save_wb(workbook, modified_sheets=[source_sheet, target_sheet])
            return {
                "source_sheet": source_sheet,
                "source_row": row_idx,
                "target_sheet": target_sheet,
                "target_row": next_row,
                "action": "moved_and_closed" if closure_details else "moved_and_consolidated"
            }
        finally:
            workbook.close()

