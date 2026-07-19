import os
import json
import webbrowser
from datetime import datetime
import tempfile
import time
import uuid
import sys
import requests
import shutil
import hashlib

# --- TRAFFIC AND SUBSCRIPTION CONTROL ---
MAX_ACTIVE_GUESTS = 5
GUEST_SESSION_TIMEOUT_MINUTES = 10
ACTIVE_GUEST_SESSIONS = {}  # Format: { session_id: last_active_timestamp }


try:
    from flask import Flask, redirect, render_template, request, send_file, url_for, session
    from werkzeug.security import generate_password_hash, check_password_hash
except ImportError as exc:
    raise RuntimeError("flask and werkzeug are required. Install them with: pip install flask werkzeug") from exc

from DC_Dashboard import DCDashboard
from DC_DataLoader import META, has_cpf_value, normalize_cpf, row_to_strings, RUNTIME_DIR, safe_text, DC_FILE, ACTIVE_CASE_SHEETS
from DC_Editor import DCEditor, split_dispatch_date
from DC_PendencyEngine import DCPendencyEngine, parse_case_date
from DC_Reports import DCReports
import sqlite3

# Try to import database manager for archived cases
try:
    from DC_DatabaseManager import db_manager
    DATABASE_AVAILABLE = True
except ImportError:
    DATABASE_AVAILABLE = False


ORG_COLUMNS = {
    "zone": "PresentZone",
    "circle": "PresentCircle",
    "division": "presentDivision",
    "office": "PresentOffice",
}

USERS_FILE = os.path.join(RUNTIME_DIR, "users.json")
VERSION_FILE = os.path.join(os.path.dirname(__file__), "version.json")

def load_version():
    """Load current version information"""
    try:
        with open(VERSION_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {
            "version": "1.0.0",
            "release_date": "2026-07-06",
            "update_server": "",
            "auto_check": False,
            "check_interval_hours": 24,
            "changelog": "Initial release"
        }

def save_version(version_data):
    """Save version information"""
    try:
        with open(VERSION_FILE, "w", encoding="utf-8") as f:
            json.dump(version_data, f, indent=4)
        return True
    except Exception:
        return False

def check_for_updates():
    """Check if updates are available from server"""
    version_data = load_version()
    if not version_data.get("update_server") or not version_data.get("auto_check"):
        return None
    
    try:
        response = requests.get(f"{version_data['update_server']}/version.json", timeout=10)
        if response.status_code == 200:
            remote_version = response.json()
            current_version = version_data["version"]
            
            # Compare versions (simple string comparison for now)
            if remote_version.get("version") != current_version:
                return {
                    "current": current_version,
                    "available": remote_version.get("version"),
                    "changelog": remote_version.get("changelog", ""),
                    "download_url": f"{version_data['update_server']}/DAMS.exe",
                    "size": remote_version.get("size", 0)
                }
    except Exception:
        pass
    
    return None

def download_update(download_url, progress_callback=None):
    """Download update file with progress tracking"""
    try:
        response = requests.get(download_url, stream=True, timeout=30)
        response.raise_for_status()
        
        total_size = int(response.headers.get('content-length', 0))
        downloaded = 0
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.exe')
        
        for chunk in response.iter_content(chunk_size=8192):
           if chunk:
                temp_file.write(chunk)
                downloaded += len(chunk)
                if progress_callback and total_size > 0:
                    progress_callback(downloaded, total_size)
        
        temp_file.close()
        return temp_file.name
    except Exception as e:
        return None

def verify_update_integrity(file_path, expected_hash=None, checksum_url=None):
    """Verify downloaded update file integrity"""
    if expected_hash:
        # Verify against provided hash
        with open(file_path, 'rb') as f:
            file_hash = hashlib.sha256(f.read()).hexdigest()
        return file_hash == expected_hash
    elif checksum_url:
        # Download and verify against server checksums
        try:
            response = requests.get(checksum_url, timeout=10)
            if response.status_code == 200:
                checksums = response.json()
                if "DAMS.exe" in checksums:
                    expected_sha256 = checksums["DAMS.exe"].get("sha256")
                    expected_size = checksums["DAMS.exe"].get("size")
                    
                    # Verify size first (faster)
                    actual_size = os.path.getsize(file_path)
                    if expected_size and actual_size != expected_size:
                        return False
                    
                    # Verify SHA256
                    if expected_sha256:
                        with open(file_path, 'rb') as f:
                            file_hash = hashlib.sha256(f.read()).hexdigest()
                        return file_hash == expected_sha256
        except Exception:
            pass
    
    # Basic file integrity check
    return os.path.exists(file_path) and os.path.getsize(file_path) > 1000000  # At least 1MB

def apply_update(update_file_path):
    """Apply the downloaded update"""
    try:
        # Get current executable path
        if getattr(sys, 'frozen', False):
            current_exe = sys.executable
        else:
            # Running from script, not exe
            return False
        
        # Create backup
        backup_path = current_exe + ".backup"
        if os.path.exists(backup_path):
            os.remove(backup_path)
        shutil.copy2(current_exe, backup_path)
        
        # Replace with new version
        shutil.copy2(update_file_path, current_exe)
        
        # Clean up temp file
        os.remove(update_file_path)
        
        return True
    except Exception as e:
        # Restore from backup if update failed
        if os.path.exists(backup_path):
            shutil.copy2(backup_path, current_exe)
        return False

def load_users():
    default_users = {
        "2266083": {
            "password": "admin",
            "name": "Nagesh D. M",
            "role": "admin",
            "jurisdiction": "All"
        },
        "rastapeth": {
            "password": "password",
            "name": "Rastapeth Circle",
            "role": "user",
            "jurisdiction": "Rastapeth Urban Circle"
        },
        "ganeshkhind": {
            "password": "password",
            "name": "Ganeshkhind Circle",
            "role": "user",
            "jurisdiction": "Ganeshkhind Urban Circle"
        },
        "punerural": {
            "password": "password",
            "name": "Pune Rural Circle",
            "role": "user",
            "jurisdiction": "Pune Rural Circle"
        }
    }
    
    if not os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "w", encoding="utf-8") as f:
                json.dump(default_users, f, indent=4)
        except Exception:
            pass
        return default_users
        
    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            users = json.load(f)
        if not users:
            users = default_users
    except Exception:
        users = default_users
        
    # Always inject guest user for public demo access
    if "guest" not in users:
        users["guest"] = {
            "password": "guest1",
            "name": "Guest Viewer",
            "role": "user",
            "jurisdiction": "All"
        }
    return users

def save_users(users):
    try:
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(users, f, indent=4)
        return True
    except Exception:
        return False

ACTIVE_SEARCH_SHEETS = {"6DC", "22DC", "23DC"}

SEARCH_HEADER_KEYWORDS = {
    "dispatch": ["dispatch", "reference", "ref", "dc no", "dc_no", "dc#", "dc number"],
    "updated": ["updated", "date", "last updated", "entry date", "report date", "updated on"],
    "remarks": ["remark", "remarks", "note", "notes", "comments"],
}


def parse_col_value_block(text):
    parsed = {}
    errors = []
    for raw_line in text.splitlines():
        line = raw_line.strip().strip(",")
        if not line:
            continue
        if "=" not in line:
            errors.append(f"Invalid input: {line}")
            continue
        key, value = line.split("=", 1)
        if not key.strip().isdigit():
            errors.append(f"Column must be numeric: {line}")
            continue
        parsed[int(key.strip())] = value.strip()
    return parsed, errors


def build_search_by_cpf(loader, editor, cpf, all_dc=None, archive_dc=None):
    emp_df = loader.load_emp()
    query_clean = str(cpf).strip()
    target_cpf = normalize_cpf(query_clean)
    
    employee = {}
    if target_cpf and not emp_df.empty and "CPFNO_NORM" in emp_df.columns:
        match = emp_df[emp_df["CPFNO_NORM"] == target_cpf]
        if not match.empty:
            employee = match.iloc[0].to_dict()
            
    cases = []
    if query_clean:
        if all_dc is None:
            all_dc = loader.load_sheets(ACTIVE_SEARCH_SHEETS, use_cache=True)
            
        if archive_dc is None:
            import os
            archive_file = os.path.join(os.path.dirname(loader.dc_file), "closed_cases_master.xlsx")
            archive_dc = {}
            if os.path.exists(archive_file):
                archive_dc = loader.load_sheets(ACTIVE_SEARCH_SHEETS, use_cache=False, filepath=archive_file)
            
        for sheet_name in sorted(ACTIVE_SEARCH_SHEETS):
            for source_tag, frame in [("Working", all_dc.get(sheet_name)), ("Archive", archive_dc.get(sheet_name))]:
                if frame is None or frame.empty:
                    continue
                meta = META.get(sheet_name, {})
                cpf_index = meta.get("cpf_col", 3) - 1
                dc_index = meta.get("dc_col", 10) - 1
                dc_record_no_index = meta.get("dc_record_no_col", 11) - 1 if "dc_record_no_col" in meta else None
                dc_record_date_index = meta.get("dc_record_date_col", 12) - 1 if "dc_record_date_col" in meta else None
                facts_index = meta.get("facts_col", 9) - 1 if "facts_col" in meta else None
                start_row = max(meta.get("data_start", 4) - 1, 0)
                
                if cpf_index >= frame.shape[1]:
                    continue
                headers = editor.get_column_labels(sheet_name)
                
                for idx, row in frame.iloc[start_row:].iterrows():
                    row_values = row_to_strings(list(row.iloc[: min(frame.shape[1], 20)]))
                    row_cpf_raw = row.iloc[cpf_index] if cpf_index < frame.shape[1] else ""
                    row_cpf = normalize_cpf(row_cpf_raw)
                    
                    cpf_match = (target_cpf and row_cpf == target_cpf)
                    
                    dc_val = str(row_values[dc_index]).strip() if dc_index < len(row_values) else ""
                    dc_match = (query_clean.lower() in dc_val.lower())
                    
                    dc_rec_val = ""
                    if dc_record_no_index is not None and dc_record_no_index < len(row_values):
                        dc_rec_val = str(row_values[dc_record_no_index]).strip()
                    dc_rec_match = (dc_rec_val and query_clean.lower() in dc_rec_val.lower())
                    
                    if cpf_match or dc_match or dc_rec_match:
                        case_fields = extract_case_fields(headers, row_values, dc_index)
                        
                        facts_of_case = ""
                        if facts_index is not None and facts_index < len(row_values):
                            facts_of_case = row_values[facts_index]
                        
                        dc_record_date = ""
                        if dc_record_date_index is not None and dc_record_date_index < len(row_values):
                            dc_record_date = row_values[dc_record_date_index]
                            
                        preview_with_headers = [
                            {
                                "column": i + 1,
                                "header": next((h.get("header", "") for h in headers if h.get("column") == i + 1), ""),
                                "value": row_values[i]
                            }
                            for i in range(min(len(row_values), 8))
                        ]
                        
                        emp_details = employee
                        if not emp_details and row_cpf:
                            if not emp_df.empty and "CPFNO_NORM" in emp_df.columns:
                                match = emp_df[emp_df["CPFNO_NORM"] == row_cpf]
                                if not match.empty:
                                    emp_details = match.iloc[0].to_dict()
                                    
                        cases.append(
                            {
                                "row_number": int(idx) + 1,
                                "sheet": sheet_name,
                                "row_data": row_values,
                                "cpfno": row_cpf_raw,
                                "employee_name": emp_details.get("EmployeeName", "") if emp_details else "",
                                "designation": emp_details.get("Designation", "") if emp_details else "",
                                "dc_no": dc_val,
                                "dc_record_no": dc_rec_val,
                                "dc_record_date": dc_record_date,
                                "facts": facts_of_case,
                                "fields": case_fields,
                                "preview": preview_with_headers,
                                "meta": meta,
                                "source": source_tag,
                                "active_case": sheet_name in ACTIVE_SEARCH_SHEETS
                            }
                        )
    cases.sort(key=lambda item: (0 if item["active_case"] else 1, item["sheet"], item["row_number"]))
    return {"employee": employee, "cases": cases, "cpfno": query_clean}


def normalize_header(text):
    if text is None:
        return ""
    return str(text).strip().lower()


def find_header_column(headers, keywords):
    for keyword in keywords:
        for item in headers:
            header_text = normalize_header(item.get("header", ""))
            if keyword in header_text:
                return item["column"] - 1
    return None


def extract_case_fields(headers, row_values, dc_index):
    dispatch_index = find_header_column(headers, SEARCH_HEADER_KEYWORDS["dispatch"])
    remarks_index = find_header_column(headers, SEARCH_HEADER_KEYWORDS["remarks"])
    
    dispatch_val = row_values[dispatch_index] if dispatch_index is not None and dispatch_index < len(row_values) else (row_values[dc_index] if dc_index < len(row_values) else "")
    _, dispatch_date = split_dispatch_date(safe_text(dispatch_val))
    
    return {
        "dispatch_ref": dispatch_val,
        "last_updated": dispatch_date,
        "remarks": row_values[remarks_index] if remarks_index is not None and remarks_index < len(row_values) else "",
    }


def build_search_by_name(loader, editor, name):
    emp_df = loader.load_emp()
    if emp_df.empty or "EmployeeName" not in emp_df.columns:
        return []
    mask = emp_df["EmployeeName"].str.lower().str.contains(str(name).strip().lower(), na=False)
    employees = emp_df[mask].head(100).to_dict("records")
    
    all_dc = loader.load_sheets(ACTIVE_SEARCH_SHEETS, use_cache=True)
    
    archive_dc = {}
    import os
    archive_file = os.path.join(os.path.dirname(loader.dc_file), "closed_cases_master.xlsx")
    if os.path.exists(archive_file):
        archive_dc = loader.load_sheets(ACTIVE_SEARCH_SHEETS, use_cache=False, filepath=archive_file)
        
    results = []
    for employee in employees:
        cpf = normalize_cpf(employee.get("CPFNO", ""))
        cases = []
        if cpf:
            cpf_results = build_search_by_cpf(loader, editor, cpf, all_dc=all_dc, archive_dc=archive_dc)
            cases = cpf_results.get("cases", [])
        results.append({"employee": employee, "cases": cases})
    return results


def get_org_values(loader, kind):
    column = ORG_COLUMNS[kind]
    emp_df = loader.load_emp()
    if emp_df.empty or column not in emp_df.columns:
        return []
    values = []
    for value in emp_df[column].fillna("").astype(str):
        text = value.strip()
        if text and text not in values:
            values.append(text)
    return sorted(values)


def collect_cases_by_org(loader, kind, value):
    column = ORG_COLUMNS[kind]
    emp_df = loader.load_emp()
    if emp_df.empty or column not in emp_df.columns:
        return []
    matched = emp_df[emp_df[column].fillna("").astype(str).str.strip().str.lower() == value.strip().lower()]
    cpf_map = {}
    for _, row in matched.iterrows():
        cpf_map[row.get("CPFNO_NORM", "")] = row.to_dict()
    if not cpf_map:
        return []

    all_dc = loader.load_sheets(loader.op_sheets, use_cache=True)
    cases = []
    for sheet_name in loader.op_sheets:
        frame = all_dc.get(sheet_name)
        if frame is None or frame.empty:
            continue
            
        meta = META.get(sheet_name, {})
        cpf_index = meta.get("cpf_col", 3) - 1
        dc_index = meta.get("dc_col", 10) - 1
        dc_record_no_index = meta.get("dc_record_no_col", 11) - 1 if "dc_record_no_col" in meta else None
        dc_record_date_index = meta.get("dc_record_date_col", 12) - 1 if "dc_record_date_col" in meta else None
        facts_index = meta.get("facts_col", 9) - 1 if "facts_col" in meta else None
        start_row = max(meta.get("data_start", 4) - 1, 0)
        if cpf_index >= frame.shape[1]:
            continue
        target_cpfs = set(cpf_map.keys())
        data_slice = frame.iloc[start_row:, cpf_index].apply(normalize_cpf)
        mask = data_slice.isin(target_cpfs)
        for idx, row in frame.iloc[start_row:][mask].iterrows():
            cpf_norm = normalize_cpf(row.iloc[cpf_index])
            employee = cpf_map.get(cpf_norm, {})
            row_values = row_to_strings(list(row.iloc[: min(frame.shape[1], 20)]))
            
            # Extract DC record no and date for sheets that have these columns
            dc_record_no = ""
            dc_record_date = ""
            if dc_record_no_index is not None and dc_record_no_index < len(row_values):
                dc_record_no = row_values[dc_record_no_index]
            if dc_record_date_index is not None and dc_record_date_index < len(row_values):
                dc_record_date = row_values[dc_record_date_index]
            
            # Extract facts of case for sheets that have this column
            facts_of_case = ""
            if facts_index is not None and facts_index < len(row_values):
                facts_of_case = row_values[facts_index]
            
            cases.append(
                {
                    "sheet": sheet_name,
                    "row_number": int(idx) + 1,
                    "cpfno": employee.get("CPFNO", ""),
                    "employee_name": employee.get("EmployeeName", ""),
                    "designation": employee.get("Designation", ""),
                    "org_value": employee.get(column, ""),
                    "dc_no": row_values[dc_index] if dc_index < len(row_values) else "",
                    "dc_record_no": dc_record_no,
                    "dc_record_date": dc_record_date,
                    "facts_of_case": facts_of_case,
                    "preview": row_values[:8],
                }
            )
    cases.sort(key=lambda item: (item["employee_name"], item["sheet"], item["row_number"]))
    return cases


def build_common_context(loader, editor, dashboard):
    return {
        "files_status": loader.ensure_source_files(),
        "sheet_choices": loader.op_sheets,
        "closed_cases": [],
        "user_role": session.get("user_role", "client"),
        "user_name": session.get("user_name", ""),
        "user_cpf": session.get("user_cpf", ""),
        "is_admin": session.get("user_role") == "admin",
        "user_jurisdiction": session.get("user_jurisdiction", "All"),
    }

def is_case_allowed(loader, row_cpf, jurisdiction):
    if jurisdiction == "All":
        return True
    emp_df = loader.load_emp(use_cache=True)
    norm_cpf = normalize_cpf(row_cpf)
    if not norm_cpf:
        return True
    if not emp_df.empty and "CPFNO_NORM" in emp_df.columns:
        match = emp_df[emp_df["CPFNO_NORM"] == norm_cpf]
        if not match.empty:
            circle = str(match.iloc[0].get("PresentCircle", "")).strip().lower()
            if circle and jurisdiction.strip().lower() not in circle:
                return False
    return True


def create_app(loader):
    from functools import wraps
    import threading

    # Asynchronously warm up caches to prevent slow first-page load
    def warm_cache():
        try:
            # Load employee master first
            loader.load_emp(use_cache=True)
            # Load key dashboard sheets
            dashboard_sheets = [
                "6DC", "22DC", "23DC", "12DC", "13DC",
                "4DC", "5DC", "20DC", "21DC", "14DC", "15DC",
                "7DC", "8DC", "24DC", "25DC"
            ]
            for sheet in dashboard_sheets:
                loader.load_dc_sheet(sheet, use_cache=True)
            # Load other operational sheets
            for sheet in loader.meta:
                if sheet not in dashboard_sheets:
                    loader.load_dc_sheet(sheet, use_cache=True)
        except Exception:
            pass

    threading.Thread(target=warm_cache, daemon=True).start()

    editor = DCEditor(loader)
    pendency_engine = DCPendencyEngine(loader)
    dashboard = DCDashboard(loader, pendency_engine)
    reports = DCReports(pendency_engine)

    import sys
    import os
    if getattr(sys, 'frozen', False):
        template_folder = os.path.join(sys._MEIPASS, 'templates')
        static_folder = os.path.join(sys._MEIPASS, 'static')
        app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
    else:
        app = Flask(__name__)
    app.config["SECRET_KEY"] = "dc-manager-offline-secure-key"
    # Configure session to work on both HTTP and HTTPS
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
    app.config["SESSION_COOKIE_SECURE"] = False
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    


    @app.before_request
    def track_guest_sessions():
        # Prune inactive sessions
        current_time = time.time()
        expired = [sid for sid, last_active in ACTIVE_GUEST_SESSIONS.items() 
                   if (current_time - last_active) > (GUEST_SESSION_TIMEOUT_MINUTES * 60)]
        for sid in expired:
            del ACTIVE_GUEST_SESSIONS[sid]

        # Update current user's timestamp if they are a guest
        if "user_cpf" in session and session.get("user_role") == "user":
            sid = session.get("guest_session_id")
            if sid:
                ACTIVE_GUEST_SESSIONS[sid] = current_time
    # ----------------------------------------

    ADMIN_CPF = "2266083"

    ADMIN_PASSWORD = "admin"
    
    def login_required(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if "user_cpf" not in session:
                return redirect("/login")
            return f(*args, **kwargs)
        return decorated_function

    def admin_required(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if "user_cpf" not in session:
                return redirect("/login")
            if session.get("user_role") != "admin":
                return redirect(url_for("home", message="Access denied. Admin privileges required.", status="error"))
            return f(*args, **kwargs)
        return decorated_function

    def render_page(template_name, message=None, status=None, **extra):
        if message is None:
            message = request.args.get("message", "")
        if status is None:
            status = request.args.get("status", "info")
            
        context = build_common_context(loader, editor, dashboard)
        context.update(extra)
        context["message"] = message
        context["status"] = status
        return render_template(template_name, **context)

    @app.get("/login")
    def login_page():
        if "user_cpf" in session:
            return redirect("/home")
        return render_template("login.html")

    @app.post("/login")
    def login():
        cpf = request.form.get("cpf", "").strip()
        password = request.form.get("password", "").strip()
        users = load_users()
        
        if cpf in users:
            user_data = users[cpf]
            
            # Check subscription / active status
            if user_data.get("status", "active") == "inactive":
                return render_template("login.html", error="Your account has been deactivated. Please contact the administrator.")
                
            # Check automated expiration date
            expires_at = user_data.get("expires_at")
            if expires_at:
                try:
                    exp_date = datetime.strptime(expires_at, "%Y-%m-%d").date()
                    if datetime.now().date() > exp_date:
                        return render_template("login.html", error="Your subscription expired on {}. Please renew to access.".format(expires_at))
                except Exception:
                    pass # Ignore malformed dates
            
            # Check guest traffic limits
            is_guest = user_data.get("role", "user") == "user"
            if is_guest:
                if len(ACTIVE_GUEST_SESSIONS) >= MAX_ACTIVE_GUESTS:
                    return render_template("login.html", error="⚠️ Server is currently at full capacity (Max users reached). Please try again in a few minutes, or contact the administrator to upgrade your tier.")
            
            stored_password = user_data["password"]
            authenticated = False
            
            if stored_password.startswith("pbkdf2:") or stored_password.startswith("scrypt:"):
                try:
                    if check_password_hash(stored_password, password):
                        authenticated = True
                except Exception:
                    pass
                    
            if stored_password == password:
                authenticated = True
                
            if authenticated:
                session["user_cpf"] = cpf
                session["user_name"] = user_data["name"]
                session["user_role"] = user_data.get("role", "user")
                session["user_jurisdiction"] = user_data.get("jurisdiction", "All")
                
                # Track guest session
                if is_guest:
                    sid = str(uuid.uuid4())
                    session["guest_session_id"] = sid
                    ACTIVE_GUEST_SESSIONS[sid] = time.time()
                    
                return redirect("/home")
                
        return render_template("login.html", error="Invalid CPF or Password")

    @app.post("/admin/clear_guests")
    @admin_required
    def clear_guest_sessions():
        count = len(ACTIVE_GUEST_SESSIONS)
        ACTIVE_GUEST_SESSIONS.clear()
        # Optionally pass a message, but we just redirect
        return redirect(request.referrer or "/home")

    @app.get("/logout")
    @login_required
    def logout():
        sid = session.get("guest_session_id")
        if sid and sid in ACTIVE_GUEST_SESSIONS:
            del ACTIVE_GUEST_SESSIONS[sid]
        session.clear()
        return redirect("/login")

    @app.get("/")
    def index():
        if "user_cpf" not in session:
            return redirect("/login")
        return redirect("/home")

    @app.get("/api/lookup-cpf/<cpf>")
    @login_required
    def lookup_cpf(cpf):
        emp = editor.get_employee_by_cpf(cpf)
        if emp:
            desg = emp.get("Designation", "")
            raw_pg = emp.get("paygrp", "")
            pg = loader.normalize_paygroup(raw_pg) or loader.get_designation_paygrp(desg) or "IV"
            seniority = loader.resolve_seniority(desg, pg)
            return {
                "success": True,
                "employee": {
                    "cpfno": emp.get("CPFNO", ""),
                    "name": emp.get("EmployeeName", ""),
                    "designation": emp.get("Designation", ""),
                    "place": emp.get("PresentOffice", ""),
                    "division": emp.get("presentDivision", ""),
                    "circle": emp.get("PresentCircle", ""),
                    "zone": emp.get("PresentZone", ""),
                    "paygroup": pg,
                    "seniority": seniority,
                    "brthdt": safe_text(emp.get("brthdt", "")),
                    "dtofretir": safe_text(emp.get("dtofretir", ""))
                }
            }
        return {"success": False, "error": "Employee not found in master list"}

    @app.get("/api/org-hierarchy")
    @login_required
    def get_org_hierarchy():
        emp_df = loader.load_emp()
        hierarchy = {}
        if not emp_df.empty:
            for _, row in emp_df.iterrows():
                z = str(row.get("PresentZone", "")).strip() or "Unknown Zone"
                c = str(row.get("PresentCircle", "")).strip() or "Unknown Circle"
                d = str(row.get("presentDivision", "")).strip() or "Unknown Division"
                o = str(row.get("PresentOffice", "")).strip() or "Unknown Location"
                
                if z not in hierarchy:
                    hierarchy[z] = {}
                if c not in hierarchy[z]:
                    hierarchy[z][c] = {}
                if d not in hierarchy[z][c]:
                    hierarchy[z][c][d] = []
                if o not in hierarchy[z][c][d] and o != "Unknown Location":
                    hierarchy[z][c][d].append(o)
        return {"success": True, "hierarchy": hierarchy}

    @app.get("/case/initiate")
    @admin_required
    def initiate_case_form():
        return render_page("initiate_case.html")

    @app.post("/case/initiate")
    @admin_required
    def initiate_case():
        cpf = request.form.get("cpf", "").strip()
        case_type = request.form.get("case_type", "").strip()
        
        vs_ref = request.form.get("vs_ref", "").strip()
        dispatch = request.form.get("dispatch", "").strip()
        facts = request.form.get("facts", "").strip()
        status = request.form.get("status", "").strip()
        remarks = request.form.get("remarks", "").strip()
        
        emp = editor.get_employee_by_cpf(cpf)
        if not emp:
            return render_page("initiate_case.html", message="Error: CPF not found in employee master.", status="error")
            
        desg = emp.get("Designation", "")
        raw_pg = emp.get("paygrp", "")
        pg = loader.normalize_paygroup(raw_pg) or loader.get_designation_paygrp(desg) or "IV"
        seniority = loader.resolve_seniority(desg, pg)
        
        dob = safe_text(emp.get("brthdt", ""))
        ret = safe_text(emp.get("dtofretir", ""))
        ret_str = f"{dob} / {ret}".strip(" /")
            
        if case_type == "minor":
            target_sheet = "4DC" if seniority == "State" else "5DC"
        elif case_type == "major":
            target_sheet = "20DC" if seniority == "State" else "21DC"
        elif case_type == "suspension":
            target_sheet = "14DC" if seniority == "State" else "15DC"
        else:
            return render_page("initiate_case.html", message="Error: Invalid case type selected.", status="error")
            
        field_map = editor.get_sheet_field_map(target_sheet)
        payload = {}
        
        def set_col(key, val):
            col = field_map.get(key)
            if col:
                payload[col] = val
                
        set_col("cpfno", cpf)
        if "designation" in field_map:
            set_col("name", emp.get("EmployeeName", ""))
            set_col("designation", emp.get("Designation", ""))
        else:
            set_col("name", f"{emp.get('EmployeeName', '')} {emp.get('Designation', '')}")
            
        set_col("paygroup", pg)
        set_col("place", emp.get("PresentOffice", ""))
        set_col("retirement", ret_str)
        set_col("theft", "Other")
        
        susp_order = request.form.get("susp_order", "").strip()
        susp_date = request.form.get("susp_date", "").strip()
        set_col("susp_order", susp_order)
        set_col("susp_date", susp_date)
        
        set_col("vs_ref", vs_ref)
        set_col("dispatch", dispatch)
        set_col("chargesheet", dispatch)
        set_col("facts", facts)
        set_col("status", status)
        set_col("remarks", remarks)
        
        try:
            res = editor.add_record(target_sheet, payload)
            loader.clear_sheet_cache(target_sheet)
            
            # Auto-upload logic
            if res and "row_number" in res:
                if target_sheet in ("4DC", "5DC"):
                    editor.copy_record_to_sheet(target_sheet, res["row_number"], "6DC")
                    loader.clear_sheet_cache("6DC")
                elif target_sheet == "20DC":
                    editor.copy_record_to_sheet(target_sheet, res["row_number"], "22DC")
                    loader.clear_sheet_cache("22DC")
                elif target_sheet == "21DC":
                    editor.copy_record_to_sheet(target_sheet, res["row_number"], "23DC")
                    loader.clear_sheet_cache("23DC")
            
            return render_page("initiate_case.html", message=f"Case successfully initiated in {target_sheet} (Row {res['row_number']}).", status="success")
        except Exception as e:
            return render_page("initiate_case.html", message=f"Error writing to Excel: {str(e)}", status="error")

    @app.post("/case/consolidate/<sheet_name>/<int:row_number>")
    @admin_required
    def consolidate_case(sheet_name, row_number):
        if sheet_name in ("4DC", "5DC"):
            target_sheet = "6DC"
        elif sheet_name == "20DC":
            target_sheet = "22DC"
        elif sheet_name == "21DC":
            target_sheet = "23DC"
        elif sheet_name == "14DC":
            target_sheet = "12DC"
        elif sheet_name == "15DC":
            target_sheet = "13DC"
        else:
            return redirect(url_for("home", message=f"Error: Sheet {sheet_name} cannot be consolidated.", status="error"))
            
        try:
            res = editor.move_or_close_case(sheet_name, row_number, target_sheet)
            return redirect(url_for("home", tab="ongoing", message=f"Case consolidated from {sheet_name} to {target_sheet} (Row {res['target_row']}).", status="success"))
        except Exception as e:
            return redirect(url_for("home", message=f"Error consolidating case: {str(e)}", status="error"))

    @app.get("/case/close/<sheet_name>/<int:row_number>")
    @admin_required
    def close_case_form(sheet_name, row_number):
        try:
            record = editor.get_record(sheet_name, row_number)
            meta = loader.get_sheet_meta(sheet_name)
            cpf_col = meta.get("cpf_col", 3)
            row_vals = record.get("row_values", [])
            cpf_val = ""
            if cpf_col <= len(row_vals):
                cpf_val = normalize_cpf(row_vals[cpf_col - 1])
            
            # Always use employee details from master
            emp = editor.get_employee_by_cpf(cpf_val) if cpf_val else None
            # If record already has employee_details from master, use that
            if record and "employee_details" in record and record["employee_details"]:
                emp = record["employee_details"]
            
            field_map = editor.get_sheet_field_map(sheet_name)
            editor_dc_col = field_map.get("dispatch") or meta.get("dc_col", 10)
            editor_facts_col = field_map.get("facts") or 9
            
            return render_page("close_case.html", 
                               record=record, 
                               employee=emp, 
                               sheet_name=sheet_name, 
                               row_number=row_number,
                               editor_dc_col=editor_dc_col,
                               editor_facts_col=editor_facts_col)
        except Exception as e:
            return redirect(url_for("home", message=str(e), status="error"))

    @app.post("/case/close/<sheet_name>/<int:row_number>")
    @admin_required
    def close_case(sheet_name, row_number):
        outcome = request.form.get("outcome", "").strip()
        punishment = request.form.get("punishment", "").strip()
        close_ref = request.form.get("close_ref", "").strip()
        close_date_raw = request.form.get("close_date", "").strip()
        
        close_date_formatted = ""
        if close_date_raw:
            try:
                dt_obj = datetime.strptime(close_date_raw, "%Y-%m-%d")
                close_date_formatted = dt_obj.strftime("%d.%m.%Y")
            except Exception:
                pass
                
        if close_date_formatted:
            full_ref = f"{close_ref} dtd {close_date_formatted}"
        else:
            full_ref = close_ref
        
        try:
            record = editor.get_record(sheet_name, row_number)
            meta = loader.get_sheet_meta(sheet_name)
            
            field_map = editor.get_sheet_field_map(sheet_name)
            pg_col = field_map.get("paygroup")
            pg_val = "III"
            if pg_col and pg_col <= len(record.get("row_values", [])):
                pg_val = str(record.get("row_values", [])[pg_col - 1]).strip()
                
            is_class_1_2 = any(c in pg_val.upper() for c in ["I", "II"]) and "III" not in pg_val.upper()
            
            if sheet_name == "6DC":
                target_sheet = "7DC" if is_class_1_2 else "8DC"
            elif sheet_name in ("22DC", "23DC"):
                target_sheet = "24DC" if is_class_1_2 else "25DC"
            else:
                if "minor" in meta.get("type", "") or sheet_name in ("4DC", "5DC"):
                    target_sheet = "7DC" if is_class_1_2 else "8DC"
                else:
                    target_sheet = "24DC" if is_class_1_2 else "25DC"
                    
            closure_details = {
                "outcome": outcome,
                "punishment": punishment,
                "close_ref": full_ref
            }
            
            res = editor.move_or_close_case(sheet_name, row_number, target_sheet, closure_details)
            return redirect(url_for("home", tab="closed", message=f"Case successfully closed and moved to {target_sheet} (Row {res['target_row']}).", status="success"))
        except Exception as e:
            return redirect(url_for("home", message=f"Error closing case: {str(e)}", status="error"))

    @app.get("/home")
    @login_required
    def home():
        message = request.args.get("message", "")
        status = request.args.get("status", "info")
        
        dashboard_data = dashboard.build_dashboard_data()
        emp_df = loader.load_emp()
        
        counts = {
            "employee_total": len(emp_df)
        }
        
        return render_page("home_new.html",
                           message=message,
                           status=status,
                           counts=counts,
                           dashboard=dashboard_data)

    @app.get("/refresh-data")
    @login_required
    def refresh_data():
        try:
            from migrate_cases import migrate_cases
            from migrate_employees import migrate_employees
            
            # Run the database migration to sync with excel
            migrate_employees()
            migrate_cases()
            
            # Clear caches if any
            loader.clear_cache()
            return redirect(url_for('home', message="Database synced and refreshed successfully from Excel files.", status="success"))
        except Exception as e:
            return redirect(url_for('home', message=f"Failed to refresh data: {str(e)}", status="error"))

    @app.get("/search/cpf")
    def search_cpf():
        cpf = request.args.get("cpfno", "").strip()
        if not cpf:
            return render_page("search_cpf.html", message="Enter a CPF number to search.", status="error", cpf_result=None)
        result = build_search_by_cpf(loader, editor, cpf)
        return render_page("search_cpf.html", message=f"Found {len(result['cases'])} case record(s) for CPF {cpf}.", status="success", cpf_result=result)

    @app.get("/search/name")
    def search_name():
        name = request.args.get("employee_name", "").strip()
        if not name:
            return render_page("search_name.html", message="Enter a name to search.", status="error", name_query="", name_result=[])
        result = build_search_by_name(loader, editor, name)
        return render_page("search_name.html", message=f"Found {len(result)} employee match(es) for '{name}'.", status="success", name_query=name, name_result=result)

    @app.get("/browse/org")
    def browse_org():
        kind = request.args.get("kind", "circle").strip().lower()
        value = request.args.get("value", "").strip()
        if kind not in ORG_COLUMNS:
            kind = "circle"
        cases = collect_cases_by_org(loader, kind, value) if value else []
        message = f"Found {len(cases)} case record(s) for {kind} '{value}'." if value else "Select zone, circle, or division to view all linked cases."
        status = "success" if value else "info"
        
        # Load org values only when browsing org page
        circle_values = get_org_values(loader, "circle")
        division_values = get_org_values(loader, "division")
        zone_values = get_org_values(loader, "zone")
        
        return render_page("org_cases.html", 
                           message=message, 
                           status=status, 
                           org_kind=kind, 
                           org_value=value, 
                           org_cases=cases,
                           circle_values=circle_values,
                           division_values=division_values,
                           zone_values=zone_values)

    @app.get("/sheets")
    def sheets_index():
        return render_page(
            "sheet_view.html",
            message="Select a sheet category and sheet to view data.",
            status="info",
            sheet_info=None,
            preview_rows=None,
        )

    @app.get("/sheet/<sheet_name>")
    def sheet_view(sheet_name):
        try:
            info = editor.view_sheet(sheet_name)
            preview_rows = editor.get_sheet_preview(sheet_name)
            
            # Reorder columns so that DC No and DC Date are next to CPF No for ALL views
            if info and preview_rows and info.get("cpf_col"):
                cpf_col = info.get("cpf_col")
                
                # Identify columns to move
                cols_to_move = []
                if info.get("dc_col"):
                    cols_to_move.append(info.get("dc_col"))       # Dispatch No
                    cols_to_move.append(info.get("dc_col") + 1)   # Dispatch Date
                if info.get("dc_record_no_col"):
                    cols_to_move.append(info.get("dc_record_no_col"))
                if info.get("dc_record_date_col"):
                    cols_to_move.append(info.get("dc_record_date_col"))
                
                headers = info.get("headers", [])
                
                # Process columns in reverse order so they stack correctly after CPF
                for move_col in reversed(cols_to_move):
                    move_idx = move_col - 1
                    cpf_idx = cpf_col - 1
                    
                    # Reorder values in preview rows
                    for row in preview_rows:
                        vals = row.get("values", [])
                        if move_idx < len(vals):
                            val = vals.pop(move_idx)
                            vals.insert(cpf_idx + 1 if cpf_idx < move_idx else cpf_idx, val)
                    
                    # Reorder headers
                    move_header = next((h for h in headers if h["column"] == move_col), None)
                    cpf_header = next((h for h in headers if h["column"] == cpf_col), None)
                    
                    if move_header and cpf_header:
                        headers.remove(move_header)
                        cpf_h_idx = headers.index(cpf_header)
                        headers.insert(cpf_h_idx + 1, move_header)
                        
                    # Adjust internal indices since we popped a column
                    if move_idx < cpf_idx:
                        cpf_col -= 1
                        
        except Exception as exc:
            return render_page("sheet_view.html", message=str(exc), status="error", sheet_info=None, preview_rows=None)
        return render_page(
            "sheet_view.html",
            message=f"Loaded structure for {sheet_name}.",
            status="success",
            sheet_info=info,
            preview_rows=preview_rows,
        )

    @app.get("/record/<sheet_name>/<int:row_number>")
    @login_required
    def record_detail(sheet_name, row_number):
        try:
            record = editor.get_record(sheet_name, row_number)
            # Ensure employee details from master are used for display
            if record and "employee_details" not in record:
                field_map = editor.get_sheet_field_map(sheet_name)
                cpf_col = field_map.get("cpfno")
                if cpf_col and cpf_col <= len(record.get("row_values", [])):
                    cpf_val = normalize_cpf(record["row_values"][cpf_col - 1])
                    if cpf_val:
                        emp = editor.get_employee_by_cpf(cpf_val)
                        record["employee_details"] = emp or {}
        except Exception as exc:
            return render_page("record_detail.html", message=str(exc), status="error", record=None)
        return render_page(
            "record_detail.html",
            message=request.args.get("message", ""),
            status=request.args.get("status", "info"),
            record=record,
        )

    @app.post("/record/<sheet_name>/<int:row_number>/update")
    @admin_required
    def update_record(sheet_name, row_number):
        # 1. Parse direct cell input fields (cell_1, cell_2, ...)
        payload = {}
        disp_nums = {}
        disp_dates = {}
        for key, value in request.form.items():
            if key.startswith("cell_dispnum_"):
                try:
                    col_num = int(key.split("_")[2])
                    disp_nums[col_num] = value.strip()
                except ValueError:
                    pass
            elif key.startswith("cell_dispdate_"):
                try:
                    col_num = int(key.split("_")[2])
                    disp_dates[col_num] = value.strip()
                except ValueError:
                    pass
            elif key.startswith("cell_"):
                try:
                    col_num = int(key.split("_")[1])
                    payload[col_num] = value.strip()
                except ValueError:
                    pass
        
        # Merge dispatch fields into payload
        for col_num in set(disp_nums.keys()).union(disp_dates.keys()):
            num = disp_nums.get(col_num, "")
            date = disp_dates.get(col_num, "")
            if num and date:
                merged = f"{num} dtd {date}"
            elif num:
                merged = num
            elif date:
                merged = f"dtd {date}"
            else:
                merged = ""
            payload[col_num] = merged
        
        # 2. Fallback to textarea block if no cell_ fields are found
        if not payload:
            payload, errors = parse_col_value_block(request.form.get("update_values", ""))
            if errors:
                return redirect(url_for("record_detail", sheet_name=sheet_name, row_number=row_number, message="; ".join(errors), status="error"))
                
        if not payload:
            return redirect(url_for("record_detail", sheet_name=sheet_name, row_number=row_number, message="No values supplied for update.", status="error"))
            
        # Validation for 6DC sheet
        if sheet_name.upper() == "6DC":
            if not str(payload.get(10, "")).strip():
                return redirect(url_for("record_detail", sheet_name=sheet_name, row_number=row_number, message="Fact of Case (Column 10) cannot be blank.", status="error"))
            if not str(payload.get(11, "")).strip():
                return redirect(url_for("record_detail", sheet_name=sheet_name, row_number=row_number, message="Dispatch No. (Column 11) cannot be blank.", status="error"))
            if not str(payload.get(19, "")).strip():
                return redirect(url_for("record_detail", sheet_name=sheet_name, row_number=row_number, message="DC No (Column 19) cannot be blank.", status="error"))
            if not str(payload.get(20, "")).strip():
                return redirect(url_for("record_detail", sheet_name=sheet_name, row_number=row_number, message="Date (Column 20) cannot be blank.", status="error"))
            
        try:
            editor.update_record(sheet_name, row_number, payload)
        except Exception as exc:
            return redirect(url_for("record_detail", sheet_name=sheet_name, row_number=row_number, message=str(exc), status="error"))
            
        return redirect(url_for("record_detail", sheet_name=sheet_name, row_number=row_number, message="Record updated successfully.", status="success"))

    @app.post("/record/<sheet_name>/<int:row_number>/delete")
    @admin_required
    def delete_record(sheet_name, row_number):
        try:
            editor.delete_record(sheet_name, row_number)
        except Exception as exc:
            return redirect(url_for("record_detail", sheet_name=sheet_name, row_number=row_number, message=str(exc), status="error"))
        return redirect(url_for("home"))

    @app.post("/extract")
    @admin_required
    def extract():
        try:
            stats = editor.extract_all_dc()
        except Exception as exc:
            return render_page("index.html", message=str(exc), status="error")
        total = sum(item["count"] for item in stats)
        return render_page("index.html", message=f"Extracted DC numbers across {len(stats)} sheets for {total} rows.", status="success")

    @app.post("/export/<fmt>")
    def export_report(fmt):
        all_dc = loader.load_all(use_cache=False)
        try:
            path = reports.export_pendency_csv(all_dc) if fmt == "csv" else reports.export_pendency_json(all_dc)
        except Exception as exc:
            return render_page("index.html", message=str(exc), status="error")
        return send_file(path, as_attachment=True, download_name=os.path.basename(path))

    @app.get("/export")
    @login_required
    def export_options():
        return render_page("export_options.html")
        
    @app.post("/archive")
    @login_required
    def run_archive():
        auto_delete = request.form.get("auto_delete") == "yes"
        try:
            archive_path = editor.archive_closed_cases(auto_delete=auto_delete)
            return send_file(archive_path, as_attachment=True, download_name="closed_cases_master.xlsx")
        except Exception as exc:
            return redirect(url_for("home", message=f"Archive Error: {str(exc)}", status="error"))

    @app.get("/download-workbook")
    @login_required
    def download_workbook():
        try:
            return send_file(loader.dc_file, as_attachment=True, download_name="origional_35_dc.xlsx")
        except Exception as exc:
            return redirect(url_for("home", message=f"Error exporting workbook: {str(exc)}", status="error"))

    @app.get("/reports/agewise")
    @login_required
    def reports_agewise():
        category = request.args.get("category", "all")
        all_dc = loader.load_sheets(loader.op_sheets, use_cache=True)
        employees = loader.load_emp()
        summary = pendency_engine.build_agewise_org_summary(all_dc, employees, category_filter=category)
        return render_page("agewise_summary.html", summary=summary, active_category=category)

    @app.get("/reports/paygroup-monthwise")
    @login_required
    def reports_paygroup_monthwise():
        category = request.args.get("category", "all")
        all_dc = loader.load_sheets(loader.op_sheets, use_cache=True)
        employees = loader.load_emp()
        summary = pendency_engine.build_paygroup_month_summary(all_dc, employees, category_filter=category)
        return render_page("paygroup_summary.html", summary=summary, active_category=category)

    @app.get("/export/agewise-excel")
    @login_required
    def export_agewise_excel():
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        all_dc = loader.load_sheets(loader.op_sheets, use_cache=True)
        employees = loader.load_emp()
        summary = pendency_engine.build_agewise_org_summary(all_dc, employees)
        
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_title = Font(name="Calibri", size=14, bold=True)
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for name, data_list in [("Initiator Office Summary", summary["circle"]), ("Paygroup Summary", summary["paygroup"])]:
            ws = wb.create_sheet(title=name)
            ws.views.sheetView[0].showGridLines = True
            
            ws.cell(1, 1, f"Age-wise Case Pendency Status - {name}").font = font_title
            ws.row_dimensions[1].height = 25
            
            headers = [
                "Name", 
                "Pending < 6 Months", 
                "Pending 6 - 12 Months", 
                "Pending 1 - 2 Years", 
                "Pending > 2 Years", 
                "Unknown Date", 
                "Total Pending"
            ]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(3, col_idx, h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_thin
            ws.row_dimensions[3].height = 24
            
            row_idx = 4
            for item in data_list:
                row_values = [
                    item["name"],
                    item["less_6m"],
                    item["six_to_twelve"],
                    item["one_to_two_yr"],
                    item["more_two_yr"],
                    item["unknown_date"],
                    item["total"]
                ]
                for col_idx, val in enumerate(row_values, 1):
                    cell = ws.cell(row_idx, col_idx, val)
                    cell.alignment = align_left if col_idx == 1 else align_right
                    cell.border = border_thin
                    if col_idx > 1:
                        cell.number_format = '#,##0'
                ws.row_dimensions[row_idx].height = 20
                row_idx += 1
                
            total_row = [
                "Grand Total",
                sum(x["less_6m"] for x in data_list),
                sum(x["six_to_twelve"] for x in data_list),
                sum(x["one_to_two_yr"] for x in data_list),
                sum(x["more_two_yr"] for x in data_list),
                sum(x["unknown_date"] for x in data_list),
                sum(x["total"] for x in data_list)
            ]
            for col_idx, val in enumerate(total_row, 1):
                cell = ws.cell(row_idx, col_idx, val)
                cell.font = Font(name="Calibri", size=11, bold=True)
                cell.alignment = align_left if col_idx == 1 else align_right
                cell.border = Border(
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='double', color='000000'),
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9')
                )
                if col_idx > 1:
                    cell.number_format = '#,##0'
            ws.row_dimensions[row_idx].height = 22
            
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.row >= 3:
                        max_len = max(max_len, len(str(cell.value or '')))
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
                
        month_name = datetime.now().strftime("%B")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_path = os.path.join(loader.cache_dir, f"{month_name}_35DC_agewise_{timestamp}.xlsx")
        wb.save(temp_path)
        return send_file(temp_path, as_attachment=True, download_name=f"{month_name}_35DC_agewise_{timestamp}.xlsx")

    @app.get("/cases/initiator")
    @login_required
    def cases_by_initiator():
        initiator = request.args.get("initiator", "")
        category = request.args.get("category", "")
        
        if not initiator or not category:
            return redirect(url_for("home", message="Missing parameters.", status="error"))
            
        all_dc = loader.load_sheets(loader.op_sheets, use_cache=True)
        active_sheets = ACTIVE_CASE_SHEETS.get(category, [])
        if not active_sheets:
            return redirect(url_for("home", message="Invalid category.", status="error"))
            
        matched_cases = []
        
        for sheet_name in active_sheets:
            frame = all_dc.get(sheet_name)
            if frame is None or frame.empty:
                continue
            meta = META.get(sheet_name, {})
            cpf_col_idx = meta.get("cpf_col", 3) - 1
            dc_col_idx = meta.get("dc_col", 10) - 1
            start_row = max(meta.get("data_start", 4) - 1, 0)
            
            headers = editor.get_column_labels(sheet_name)
            
            for idx, row in frame.iloc[start_row:].iterrows():
                cpf = row.iloc[cpf_col_idx] if cpf_col_idx < frame.shape[1] else ""
                if not has_cpf_value(cpf):
                    continue
                
                initiator_circle = "N.A."
                for val in reversed(row.tolist()):
                    text = str(val).upper().strip()
                    if not text: continue
                    if 'RPUC' in text or 'RASTAPETH' in text or 'RASTA PETH' in text:
                        initiator_circle = "Rastapeth Urban Circle"
                        break
                    elif 'GKUC' in text or 'GANESHKHIND' in text:
                        initiator_circle = "Ganeshkhind Urban Circle"
                        break
                    elif 'PRC' in text or 'PUNE RURAL' in text:
                        initiator_circle = "Pune Rural Circle"
                        break
                    elif 'PZ' in text or 'PUNE ZONE' in text:
                        initiator_circle = "Pune Zone"
                        break
                
                if initiator_circle == "N.A.":
                    dispatch_val = str(row.iloc[dc_col_idx]).upper() if dc_col_idx < frame.shape[1] else ""
                    if 'RPUC' in dispatch_val or 'RASTAPETH' in dispatch_val or 'RASTA PETH' in dispatch_val:
                        initiator_circle = "Rastapeth Urban Circle"
                    elif 'GKUC' in dispatch_val or 'GANESHKHIND' in dispatch_val:
                        initiator_circle = "Ganeshkhind Urban Circle"
                    elif 'PRC' in dispatch_val or 'PUNE RURAL' in dispatch_val:
                        initiator_circle = "Pune Rural Circle"
                    elif 'PZ' in dispatch_val or 'PUNE ZONE' in dispatch_val:
                        initiator_circle = "Pune Zone"
                    else:
                        initiator_circle = "Other / Unknown Initiator"
                        
                if initiator_circle == initiator:
                    dispatch_val = str(row.iloc[dc_col_idx]) if dc_col_idx < frame.shape[1] else ""
                    _, date_str = split_dispatch_date(safe_text(dispatch_val))
                    facts_col_idx = meta.get("facts_col", 0) - 1
                    fact_val = str(row.iloc[facts_col_idx]) if facts_col_idx >= 0 and facts_col_idx < frame.shape[1] else ""
                    
                    emp = editor.get_employee_by_cpf(normalize_cpf(cpf))
                    emp_name = emp.get("EmployeeName", "") if emp else ""
                    emp_desg = emp.get("Designation", "") if emp else ""
                    emp_place = emp.get("PresentOffice", "") if emp else ""
                    
                    if not emp_name:
                        name_col_idx = meta.get("name_col", 2) - 1
                        row_values = row_to_strings(list(row.iloc[: min(frame.shape[1], 20)]))
                        if name_col_idx < len(row_values):
                            emp_name = row_values[name_col_idx]
                            
                    matched_cases.append({
                        "sheet": sheet_name,
                        "row_number": int(idx) + 1,
                        "cpf": cpf,
                        "name": emp_name,
                        "designation": emp_desg,
                        "fact_of_case": fact_val,
                        "dc_no": dispatch_val,
                        "dc_date": date_str,
                        "place": emp_place,
                        "dc_ref": dispatch_val,
                        "dispatch_date": date_str
                    })
                    
        matched_cases.sort(key=lambda x: (x["sheet"], x["row_number"]))
        
        title = f"{category.title()} Cases for {initiator}"
        
        return render_page("cases_by_date.html", 
                           cases=matched_cases, 
                           title=title, 
                           group_type="initiator")

    @app.get("/cases/date")
    @login_required
    def cases_by_date():
        group_type = request.args.get("type", "monthly")
        year_str = request.args.get("year", "")
        month_str = request.args.get("month", "")
        
        try:
            year = int(year_str) if year_str else None
            month = int(month_str) if month_str else None
        except ValueError:
            return redirect(url_for("home", message="Invalid year or month parameter.", status="error"))
            
        all_dc = loader.load_sheets(loader.op_sheets, use_cache=True)
        active_sheets = ["6DC", "22DC", "23DC"]
        
        matched_cases = []
        
        for sheet_name in active_sheets:
            frame = all_dc.get(sheet_name)
            if frame is None or frame.empty:
                continue
            meta = META.get(sheet_name, {})
            cpf_col_idx = meta.get("cpf_col", 3) - 1
            dc_col_idx = meta.get("dc_col", 10) - 1
            start_row = max(meta.get("data_start", 4) - 1, 0)
            
            headers = editor.get_column_labels(sheet_name)
            
            for idx, row in frame.iloc[start_row:].iterrows():
                cpf = row.iloc[cpf_col_idx] if cpf_col_idx < frame.shape[1] else ""
                if not has_cpf_value(cpf):
                    continue
                
                dispatch_val = row.iloc[dc_col_idx] if dc_col_idx < frame.shape[1] else ""
                _, date_str = split_dispatch_date(safe_text(dispatch_val))
                facts_col_idx = meta.get("facts_col", 0) - 1
                fact_val = str(row.iloc[facts_col_idx]) if facts_col_idx >= 0 and facts_col_idx < frame.shape[1] else ""
                dt = parse_case_date(date_str)
                
                if dt:
                    match_year = (dt.year == year)
                    match_month = (dt.month == month) if month else True
                    
                    if match_year and match_month:
                        row_values = row_to_strings(list(row.iloc[: min(frame.shape[1], 20)]))
                        case_fields = extract_case_fields(headers, row_values, dc_col_idx)
                        
                        emp = editor.get_employee_by_cpf(normalize_cpf(cpf))
                        emp_name = emp.get("EmployeeName", "") if emp else ""
                        emp_desg = emp.get("Designation", "") if emp else ""
                        emp_place = emp.get("PresentOffice", "") if emp else ""
                        
                        if not emp_name:
                            name_col_idx = meta.get("name_col", 2) - 1
                            if name_col_idx < len(row_values):
                                emp_name = row_values[name_col_idx]
                                
                        matched_cases.append({
                            "sheet": sheet_name,
                            "row_number": int(idx) + 1,
                            "cpf": cpf,
                            "name": emp_name,
                            "designation": emp_desg,
                            "fact_of_case": fact_val,
                            "dc_no": dispatch_val,
                            "dc_date": date_str,
                            "place": emp_place,
                            "dc_ref": dispatch_val,
                            "dispatch_date": date_str
                        })
                        
        matched_cases.sort(key=lambda x: (x["sheet"], x["row_number"]))
        
        title = f"Cases from {datetime(year, month, 1).strftime('%B %Y')}" if group_type == "monthly" and month else f"Cases from Year {year}"
        
        return render_page("cases_by_date.html", 
                           cases=matched_cases, 
                           title=title, 
                           group_type=group_type, 
                           year=year, 
                           month=month)

    @app.route("/cases/agewise")
    # @login_required
    def cases_agewise():
        org_kind = request.args.get("org_kind", "")
        org_value = request.args.get("org_value", "")
        age_group = request.args.get("age_group", "")
        category = request.args.get("category", "all")
        
        all_dc = loader.load_sheets(loader.op_sheets, use_cache=True)
        if category == "all":
            active_sheets = ACTIVE_CASE_SHEETS.get("minor", []) + ACTIVE_CASE_SHEETS.get("major", [])
        else:
            active_sheets = ACTIVE_CASE_SHEETS.get(category, [])
        
        employees = loader.load_emp()
        employee_lookup = {}
        if not employees.empty:
            for _, row in employees.iterrows():
                employee_lookup[row.get("CPFNO_NORM", "")] = {
                    "zone": safe_text(row.get("PresentZone")) or "Unknown",
                    "circle": safe_text(row.get("PresentCircle")) or "Unknown",
                    "division": safe_text(row.get("presentDivision")) or "Unknown",
                    "office": safe_text(row.get("PresentOffice")) or "Unknown",
                    "PayGroup": row.get("PayGroup", row.get("paygrp", "Unknown")),
                    "EmployeeName": row.get("EmployeeName", ""),
                }
                
        matched_cases = []
        ref_date = datetime.now()
        
        for sheet_name in active_sheets:
            frame = all_dc.get(sheet_name)
            if frame is None or frame.empty:
                continue
            meta = META.get(sheet_name, {})
            cpf_col_idx = meta.get("cpf_col", 3) - 1
            dc_col_idx = meta.get("dc_col", 10) - 1
            start_row = max(meta.get("data_start", 4) - 1, 0)
            
            headers = editor.get_column_labels(sheet_name)
            
            for idx, row in frame.iloc[start_row:].iterrows():
                cpf_raw = row.iloc[cpf_col_idx] if cpf_col_idx < frame.shape[1] else ""
                cpf = normalize_cpf(cpf_raw)
                if not has_cpf_value(cpf):
                    continue
                    
                org = employee_lookup.get(cpf, {})
                
                initiator_circle = "N.A."
                for val in reversed(row.tolist()):
                    text = str(val).upper().strip()
                    if not text: continue
                    if 'RPUC' in text or 'RASTAPETH' in text or 'RASTA PETH' in text:
                        initiator_circle = "Rastapeth Urban Circle"
                        break
                    elif 'GKUC' in text or 'GANESHKHIND' in text:
                        initiator_circle = "Ganeshkhind Urban Circle"
                        break
                    elif 'PRC' in text or 'PUNE RURAL' in text:
                        initiator_circle = "Pune Rural Circle"
                        break
                    elif 'PZ' in text or 'PUNE ZONE' in text:
                        initiator_circle = "Pune Zone"
                        break
                
                if initiator_circle == "N.A.":
                    dispatch_val = str(row.iloc[dc_col_idx]).upper() if dc_col_idx < frame.shape[1] else ""
                    if 'RPUC' in dispatch_val or 'RASTAPETH' in dispatch_val or 'RASTA PETH' in dispatch_val:
                        initiator_circle = "Rastapeth Urban Circle"
                    elif 'GKUC' in dispatch_val or 'GANESHKHIND' in dispatch_val:
                        initiator_circle = "Ganeshkhind Urban Circle"
                    elif 'PRC' in dispatch_val or 'PUNE RURAL' in dispatch_val:
                        initiator_circle = "Pune Rural Circle"
                    elif 'PZ' in dispatch_val or 'PUNE ZONE' in dispatch_val:
                        initiator_circle = "Pune Zone"
                    else:
                        initiator_circle = "Other / Unknown Initiator"
                        
                pg_raw = org.get("PayGroup", org.get("paygrp", "Unknown"))
                if not pg_raw or pg_raw == "Unknown":
                    if frame.shape[1] > 4:
                        fallback_pg = str(row.iloc[4]).strip()
                        if fallback_pg and fallback_pg.lower() not in ("nan", "none", "unknown"):
                            pg_raw = fallback_pg

                if pg_raw in ("1", 1.0, "I", "1.0", 1): pg = "Pay Group I"
                elif pg_raw in ("2", 2.0, "II", "2.0", 2): pg = "Pay Group II"
                elif pg_raw in ("3", 3.0, "III", "3.0", 3): 
                    is_state = False
                    if sheet_name in ("22DC", "12DC", "4DC"):
                        is_state = True
                    elif sheet_name not in ("23DC", "13DC", "5DC"):
                        if frame.shape[1] > 4:
                            col4 = str(row.iloc[4]).upper().strip()
                            if "-S" in col4 or "STATE" in col4:
                                is_state = True
                    pg = "Pay Group III (State)" if is_state else "Pay Group III (Circle)"
                elif pg_raw in ("4", 4.0, "IV", "4.0", 4): pg = "Pay Group IV"
                else: pg = "Unknown"

                if org_kind == "circle" and initiator_circle != org_value:
                    continue
                if org_kind == "paygroup" and pg != org_value:
                    continue
                
                dispatch_val = row.iloc[dc_col_idx] if dc_col_idx < frame.shape[1] else ""
                _, date_str = split_dispatch_date(safe_text(dispatch_val))
                facts_col_idx = meta.get("facts_col", 0) - 1
                fact_val = str(row.iloc[facts_col_idx]) if facts_col_idx >= 0 and facts_col_idx < frame.shape[1] else ""
                dt = parse_case_date(date_str)
                
                category = "unknown_date"
                if dt:
                    diff_days = (ref_date - dt).days
                    if diff_days < 183:
                        category = "less_6m"
                    elif diff_days < 365:
                        category = "six_to_twelve"
                    elif diff_days < 730:
                        category = "one_to_two_yr"
                    else:
                        category = "more_two_yr"
                
                if not age_group or age_group == "all" or category == age_group:
                    row_values = row_to_strings(list(row.iloc[: min(frame.shape[1], 20)]))
                    emp_name = org.get("EmployeeName", "")
                    if not emp_name:
                        name_col_idx = meta.get("name_col", 2) - 1
                        if name_col_idx < len(row_values):
                            emp_name = row_values[name_col_idx]
                            
                    matched_cases.append({
                        "sheet": sheet_name,
                        "row_number": int(idx) + 1,
                        "cpf": cpf_raw,
                        "name": emp_name,
                        "designation": "",
                        "fact_of_case": fact_val,
                        "dc_no": dispatch_val,
                        "dc_date": date_str,
                        "place": "",
                        "dc_ref": dispatch_val,
                        "dispatch_date": date_str
                    })
                    
        matched_cases.sort(key=lambda x: (x["sheet"], x["row_number"]))
        
        if not age_group or age_group == "all":
            title = f"All Pending Cases for {org_value}"
        else:
            age_labels = {
                "less_6m": "< 6 Months",
                "six_to_twelve": "6 to 12 Months",
                "one_to_two_yr": "1 to 2 Years",
                "more_two_yr": "> 2 Years",
                "unknown_date": "Unknown Date"
            }
            title = f"{org_value} Cases pending {age_labels.get(age_group, age_group)}"
        
        return render_page("cases_by_date.html", 
                           cases=matched_cases, 
                           title=title, 
                           group_type="agewise", 
                           year=None, 
                           month=None)

    @app.get("/archive")
    @login_required
    def archive_page():
        """Archived cases page"""
        if not DATABASE_AVAILABLE:
            return render_page("error.html", 
                             title="Database Not Available",
                             message="Archived cases functionality requires database.")
        
        # Get search parameters
        search_term = request.args.get("search", "").strip()
        cpf_no = request.args.get("cpf", "").strip()
        employee_name = request.args.get("name", "").strip()
        case_type = request.args.get("type", "").strip()
        archive_month = request.args.get("month", "").strip()
        
        # Search archived cases
        results = db_manager.search_archived_cases(
            search_term=search_term if search_term else None,
            cpf_no=cpf_no if cpf_no else None,
            employee_name=employee_name if employee_name else None,
            case_type=case_type if case_type else None,
            archive_month=archive_month if archive_month else None,
            limit=100
        )
        
        # Get available archive months
        with db_manager.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT archive_month, COUNT(*) as count
                FROM archived_cases 
                GROUP BY archive_month 
                ORDER BY archive_month DESC
            """)
            months = [{"month": row[0], "count": row[1]} for row in cursor.fetchall()]
        
        return render_page("archive.html", 
                         cases=results,
                         months=months,
                         search_term=search_term,
                         cpf_no=cpf_no,
                         employee_name=employee_name,
                         case_type=case_type,
                         archive_month=archive_month)

    @app.get("/archive/<int:case_id>")
    @login_required
    def archive_case_detail(case_id):
        """View details of a specific archived case"""
        if not DATABASE_AVAILABLE:
            return render_page("error.html",
                             title="Database Not Available",
                             message="Archived cases functionality requires database.")
        
        case = db_manager.get_archived_case_by_id(case_id)
        if not case:
            return render_page("error.html",
                             title="Case Not Found",
                             message="The requested archived case was not found.")
        
        return render_page("archive_detail.html", case=case)

    @app.post("/archive/process")
    @admin_required
    def process_archive():
        """Process archival of previous month's closed cases"""
        if not DATABASE_AVAILABLE:
            return redirect(url_for("archive_page", message="Database not available", status="error"))
        
        from datetime import datetime
        
        # Get previous month
        today = datetime.now()
        if today.month == 1:
            prev_month = today.replace(year=today.year - 1, month=12)
        else:
            prev_month = today.replace(month=today.month - 1)
        
        target_month = prev_month.strftime("%Y-%m")
        
        archived_count = db_manager.archive_closed_cases(target_month, archived_by=session.get("user_cpf", "admin"))
        
        return redirect(url_for("archive_page", 
                               message=f"Archived {archived_count} cases from {target_month}",
                               status="success"))

    @app.get("/users")
    @admin_required
    def manage_users():
        users = load_users()
        return render_page("users.html", users=users)

    @app.post("/users/add")
    @admin_required
    def add_user():
        cpf = request.form.get("cpf", "").strip()
        password = request.form.get("password", "").strip()
        name = request.form.get("name", "").strip()
        role = request.form.get("role", "").strip()
        
        if not cpf or not password or not name:
            return redirect(url_for("manage_users", message="CPF, Password and Name are required.", status="error"))
            
        # Only the main admin (2266083) can create other admin users
        if role == "admin" and session.get("user_cpf") != "2266083":
            return redirect(url_for("manage_users", message="Only the main administrator can create admin users.", status="error"))
            
        users = load_users()
        if cpf in users:
            return redirect(url_for("manage_users", message="User with this CPF already exists.", status="error"))
            
        users[cpf] = {
            "password": password,
            "name": name,
            "role": role or "client"
        }
        save_users(users)
        return redirect(url_for("manage_users", message="User added successfully.", status="success"))

    @app.post("/users/update-password")
    @login_required
    def update_password():
        cpf = request.form.get("cpf", "").strip()
        new_password = request.form.get("password", "").strip()
        
        if session.get("user_cpf") != "2266083" and session.get("user_cpf") != cpf:
            return redirect(url_for("manage_users", message="You can only change your own password.", status="error"))
        
        if not cpf or not new_password:
            return redirect(url_for("manage_users", message="CPF and New Password are required.", status="error"))
            
        users = load_users()
        if cpf not in users:
            return redirect(url_for("manage_users", message="User not found.", status="error"))
            
        users[cpf]["password"] = new_password
        save_users(users)
        return redirect(url_for("manage_users", message=f"Password updated successfully for user {cpf}.", status="success"))

    @app.post("/users/delete/<cpf>")
    @admin_required
    def delete_user(cpf):
            
        if cpf == "2266083":
            return redirect(url_for("manage_users", message="Cannot delete the default administrator account.", status="error"))
            
        users = load_users()
        if cpf in users:
            del users[cpf]
            save_users(users)
            return redirect(url_for("manage_users", message="User deleted successfully.", status="success"))
        return redirect(url_for("manage_users", message="User not found.", status="error"))

    # --- AUTO UPDATE SYSTEM ---
    @app.get("/updates")
    @admin_required
    def updates_page():
        version_data = load_version()
        update_info = check_for_updates()
        return render_page("updates.html", version=version_data, update_info=update_info)

    @app.post("/updates/check")
    @admin_required
    def check_updates():
        update_info = check_for_updates()
        if update_info:
            return {
                "success": True,
                "has_update": True,
                "current": update_info["current"],
                "available": update_info["available"],
                "changelog": update_info["changelog"],
                "size": update_info["size"]
            }
        else:
            return {
                "success": True,
                "has_update": False,
                "message": "No updates available or update server not configured"
            }

    @app.post("/updates/download")
    @admin_required
    def download_update():
        update_info = check_for_updates()
        if not update_info:
            return {"success": False, "error": "No updates available"}
        
        def progress_callback(downloaded, total):
            # Could implement WebSocket or similar for real-time progress
            pass
        
        version_data = load_version()
        checksum_url = f"{version_data['update_server']}/checksums.json" if version_data.get("update_server") else None
        
        update_file = download_update(update_info["download_url"], progress_callback)
        if update_file and verify_update_integrity(update_file, checksum_url=checksum_url):
            return {
                "success": True,
                "file_path": update_file,
                "message": "Update downloaded and verified successfully"
            }
        else:
            return {"success": False, "error": "Failed to download or verify update"}

    @app.post("/updates/apply")
    @admin_required
    def apply_update_route():
        file_path = request.form.get("file_path")
        if not file_path or not os.path.exists(file_path):
            return {"success": False, "error": "Invalid update file"}
        
        if apply_update(file_path):
            return {"success": True, "message": "Update applied successfully. Please restart the application."}
        else:
            return {"success": False, "error": "Failed to apply update"}

    @app.post("/updates/configure")
    @admin_required
    def configure_updates():
        version_data = load_version()
        version_data["update_server"] = request.form.get("update_server", "").strip()
        version_data["auto_check"] = request.form.get("auto_check") == "true"
        version_data["check_interval_hours"] = int(request.form.get("check_interval_hours", 24))
        
        if save_version(version_data):
            return redirect(url_for("updates_page", message="Update configuration saved successfully.", status="success"))
        else:
            return redirect(url_for("updates_page", message="Failed to save update configuration.", status="error"))

    @app.post("/exit")
    @login_required
    def exit_app():
        import signal
        import threading

        # Clear session and delete cookie before shutdown
        session.clear()
        response = "<h3>DAMS Program Closed Successfully.</h3><p>The DAMS standalone server has been shut down. You can now close this browser window/tab.</p>"

        def shutdown():
            import time
            time.sleep(0.5)
            try:
                os.kill(os.getpid(), signal.SIGINT)
            except Exception:
                os._exit(0)

        threading.Thread(target=shutdown).start()
        return response

    @app.get("/open")
    def open_home():
        return redirect(url_for("home"))

    @app.get("/import")
    @admin_required
    def import_page():
        return render_page("import.html")

    @app.post("/import")
    @admin_required
    def import_excel():
        if "file" not in request.files:
            return redirect(url_for("import_page", message="No file uploaded", status="error"))

        file = request.files["file"]
        if file.filename == "":
            return redirect(url_for("import_page", message="No file selected", status="error"))

        if not file.filename.endswith((".xlsx", ".xls")):
            return redirect(url_for("import_page", message="Invalid file format. Please upload Excel file.", status="error"))

        try:
            import pandas as pd
            import openpyxl

            # Save uploaded file temporarily
            temp_path = os.path.join(tempfile.gettempdir(), f"import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            file.save(temp_path)

            # Load the uploaded Excel
            uploaded_wb = openpyxl.load_workbook(temp_path)
            available_sheets = uploaded_wb.sheetnames

            # Check for required sheets
            required_sheets = ["4DC", "5DC", "20DC", "21DC"]
            missing_sheets = [s for s in required_sheets if s not in available_sheets]

            if missing_sheets:
                os.remove(temp_path)
                return redirect(url_for("import_page", message=f"Missing required sheets: {', '.join(missing_sheets)}", status="error"))

            # Create backup
            backup_manager = DCBackupManager()
            backup_path = backup_manager.create_backup()

            # Load current data
            loader.load_all(use_cache=False)

            # Clear data from source sheets and import new data
            import_summary = {
                "cleared": {},
                "imported": {},
                "added_to_ongoing": {},
                "invalid_cpf": []
            }

            # Process minor cases (4DC, 5DC → 6DC)
            for sheet_name in ["4DC", "5DC"]:
                meta = META[sheet_name]
                sheet = loader.workbook[sheet_name]
                data_start = meta["data_start"]

                # Clear existing data (keep headers)
                cleared_count = 0
                for row_idx in range(data_start, sheet.max_row + 1):
                    for col in sheet[row_idx]:
                        col.value = None
                    cleared_count += 1
                import_summary["cleared"][sheet_name] = cleared_count

                # Import new data from uploaded file
                uploaded_sheet = uploaded_wb[sheet_name]
                imported_count = 0
                for row_idx, row in enumerate(uploaded_sheet.iter_rows(min_row=data_start), start=data_start):
                    for col_idx, cell in enumerate(row, start=1):
                        sheet.cell(row_idx, col_idx, cell.value)
                    imported_count += 1
                import_summary["imported"][sheet_name] = imported_count

            # Process major cases (20DC, 21DC → 22DC, 23DC)
            for sheet_name in ["20DC", "21DC"]:
                meta = META[sheet_name]
                sheet = loader.workbook[sheet_name]
                data_start = meta["data_start"]

                # Clear existing data (keep headers)
                cleared_count = 0
                for row_idx in range(data_start, sheet.max_row + 1):
                    for col in sheet[row_idx]:
                        col.value = None
                    cleared_count += 1
                import_summary["cleared"][sheet_name] = cleared_count

                # Import new data from uploaded file
                uploaded_sheet = uploaded_wb[sheet_name]
                imported_count = 0
                for row_idx, row in enumerate(uploaded_sheet.iter_rows(min_row=data_start), start=data_start):
                    for col_idx, cell in enumerate(row, start=1):
                        sheet.cell(row_idx, col_idx, cell.value)
                    imported_count += 1
                import_summary["imported"][sheet_name] = imported_count

            # Add new cases to ongoing sheets (6DC, 22DC, 23DC)
            # For minor: 4DC, 5DC → 6DC
            ongoing_6dc = loader.workbook["6DC"]
            data_start_6dc = META["6DC"]["data_start"]
            cpf_col_6dc = META["6DC"]["cpf_col"]

            for source_sheet in ["4DC", "5DC"]:
                source_data = loader.workbook[source_sheet]
                data_start_source = META[source_sheet]["data_start"]
                cpf_col_source = META[source_sheet]["cpf_col"]

                added_count = 0
                for row_idx in range(data_start_source, source_data.max_row + 1):
                    cpf_cell = source_data.cell(row_idx, cpf_col_source)
                    if cpf_cell.value:
                        # Check if CPF already exists in 6DC to avoid duplicates
                        cpf_value = normalize_cpf(str(cpf_cell.value))
                        exists = False
                        for check_row in range(data_start_6dc, ongoing_6dc.max_row + 1):
                            check_cpf = normalize_cpf(str(ongoing_6dc.cell(check_row, cpf_col_6dc).value))
                            if check_cpf == cpf_value:
                                exists = True
                                break

                        if not exists:
                            # Copy row to 6DC
                            for col_idx in range(1, source_data.max_column + 1):
                                ongoing_6dc.cell(ongoing_6dc.max_row + 1, col_idx, source_data.cell(row_idx, col_idx).value)
                            added_count += 1

                import_summary["added_to_ongoing"][f"{source_sheet}→6DC"] = added_count

            # For major: 20DC → 22DC, 21DC → 23DC
            for source_sheet, target_sheet in [("20DC", "22DC"), ("21DC", "23DC")]:
                source_data = loader.workbook[source_sheet]
                target_data = loader.workbook[target_sheet]
                data_start_source = META[source_sheet]["data_start"]
                data_start_target = META[target_sheet]["data_start"]
                cpf_col_source = META[source_sheet]["cpf_col"]
                cpf_col_target = META[target_sheet]["cpf_col"]

                added_count = 0
                for row_idx in range(data_start_source, source_data.max_row + 1):
                    cpf_cell = source_data.cell(row_idx, cpf_col_source)
                    if cpf_cell.value:
                        # Check if CPF already exists in target sheet to avoid duplicates
                        cpf_value = normalize_cpf(str(cpf_cell.value))
                        exists = False
                        for check_row in range(data_start_target, target_data.max_row + 1):
                            check_cpf = normalize_cpf(str(target_data.cell(check_row, cpf_col_target).value))
                            if check_cpf == cpf_value:
                                exists = True
                                break

                        if not exists:
                            # Copy row to target sheet
                            for col_idx in range(1, source_data.max_column + 1):
                                target_data.cell(target_data.max_row + 1, col_idx, source_data.cell(row_idx, col_idx).value)
                            added_count += 1

                import_summary["added_to_ongoing"][f"{source_sheet}→{target_sheet}"] = added_count

            # Save changes
            loader.workbook.save(DC_FILE)
            loader.clear_cache()

            # Clean up temp file
            os.remove(temp_path)

            return redirect(url_for("import_page", message=f"Import successful! {import_summary}", status="success"))

        except Exception as e:
            return redirect(url_for("import_page", message=f"Import failed: {str(e)}", status="error"))

    @app.get("/eo-reviews")
    @login_required
    def eo_reviews():
        sheets = ["22DC", "23DC"]
        all_dc = loader.load_sheets(sheets, use_cache=True)
        
        report_pending_cases = []
        scn_pending_cases = []
        scn_issued_cases = []
        not_appointed = []
        
        summary = {
            "total": {"class_1_2": 0, "class_3_4": 0, "all": 0},
            "not_appointed": {"class_1_2": 0, "class_3_4": 0, "all": 0},
            "report_pending": {"class_1_2": 0, "class_3_4": 0, "all": 0},
            "scn_pending": {"class_1_2": 0, "class_3_4": 0, "all": 0},
            "scn_issued": {"class_1_2": 0, "class_3_4": 0, "all": 0}
        }
        
        for sheet_name in sheets:
            frame = all_dc.get(sheet_name)
            if frame is None or frame.empty:
                continue
            meta = META.get(sheet_name, {})
            cpf_index = meta.get("cpf_col", 4) - 1
            dc_index = meta.get("dc_col", 10) - 1
            # EO Appt Details is Col 13 -> index 12
            eo_index = 12
            # Date of receipt of enquiry findings is Col 14 -> index 13
            report_index = 13
            # SCN Date is Col 15 -> index 14
            scn_index = 14
            start_row = max(meta.get("data_start", 5) - 1, 0)
            
            for idx, row in frame.iloc[start_row:].iterrows():
                row_values = row_to_strings(list(row.iloc[: min(frame.shape[1], 30)]))
                if not any(v.strip() for v in row_values):
                    continue
                
                cpf = row_values[cpf_index] if cpf_index < len(row_values) else ""
                dispatch_val = row_values[dc_index] if dc_index < len(row_values) else ""
                dc_no_val, dc_date_val = split_dispatch_date(safe_text(dispatch_val))
                name_designation = row_values[1] if 1 < len(row_values) else "" # Col 2
                
                eo_detail = str(row_values[eo_index]).strip() if eo_index < len(row_values) else ""
                report_detail = str(row_values[report_index]).strip() if report_index < len(row_values) else ""
                report_received = bool(report_detail and report_detail.lower() not in ["", "nan", "none"])
                
                scn_detail = str(row_values[scn_index]).strip() if scn_index < len(row_values) else ""
                scn_issued = bool(scn_detail and scn_detail.lower() not in ["", "nan", "none"])
                
                case_info = {
                    "sheet": sheet_name,
                    "row_number": int(idx) + 1,
                    "cpf": cpf,
                    "name": name_designation,
                    "dc_no": dc_no_val,
                    "dc_date": dc_date_val,
                    "eo_detail": eo_detail,
                    "report_detail": report_detail,
                    "report_received": report_received,
                    "scn_detail": scn_detail,
                    "scn_issued": scn_issued
                }
                
                cls_key = "class_1_2" if sheet_name == "22DC" else "class_3_4"
                summary["total"][cls_key] += 1
                summary["total"]["all"] += 1
                
                if eo_detail and eo_detail.lower() not in ["", "nan", "none"]:
                    if report_received:
                        if scn_issued:
                            summary["scn_issued"][cls_key] += 1
                            summary["scn_issued"]["all"] += 1
                            scn_issued_cases.append(case_info)
                        else:
                            summary["scn_pending"][cls_key] += 1
                            summary["scn_pending"]["all"] += 1
                            scn_pending_cases.append(case_info)
                    else:
                        summary["report_pending"][cls_key] += 1
                        summary["report_pending"]["all"] += 1
                        report_pending_cases.append(case_info)
                else:
                    not_appointed.append(case_info)
                    summary["not_appointed"][cls_key] += 1
                    summary["not_appointed"]["all"] += 1
                    
        return render_page("eo_reviews.html", 
                           report_pending=report_pending_cases, 
                           scn_pending=scn_pending_cases, 
                           scn_issued=scn_issued_cases, 
                           not_appointed=not_appointed, 
                           summary=summary)

    @app.get("/export/eo-reviews")
    @login_required
    def export_eo_reviews():
        sheets = ["22DC", "23DC"]
        all_dc = loader.load_sheets(sheets, use_cache=True)
        
        import csv
        import io
        from flask import Response
        
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(["Status", "Type", "Sheet", "Row", "CPF", "Name & Designation", "DC No", "EO Details", "Report Status", "Report Details", "SCN Status", "SCN Details"])
        
        for sheet_name in sheets:
            frame = all_dc.get(sheet_name)
            if frame is None or frame.empty:
                continue
            meta = META.get(sheet_name, {})
            cpf_index = meta.get("cpf_col", 4) - 1
            dc_index = meta.get("dc_col", 10) - 1
            eo_index = 12
            report_index = 13
            scn_index = 14
            start_row = max(meta.get("data_start", 5) - 1, 0)
            
            for idx, row in frame.iloc[start_row:].iterrows():
                row_values = row_to_strings(list(row.iloc[: min(frame.shape[1], 30)]))
                if not any(v.strip() for v in row_values):
                    continue
                
                cpf = row_values[cpf_index] if cpf_index < len(row_values) else ""
                dispatch_val = row_values[dc_index] if dc_index < len(row_values) else ""
                dc_no_val, dc_date_val = split_dispatch_date(safe_text(dispatch_val))
                name_designation = row_values[1] if 1 < len(row_values) else ""
                eo_detail = str(row_values[eo_index]).strip() if eo_index < len(row_values) else ""
                report_detail = str(row_values[report_index]).strip() if report_index < len(row_values) else ""
                report_received = bool(report_detail and report_detail.lower() not in ["", "nan", "none"])
                
                scn_detail = str(row_values[scn_index]).strip() if scn_index < len(row_values) else ""
                scn_issued = bool(scn_detail and scn_detail.lower() not in ["", "nan", "none"])
                
                if eo_detail and eo_detail.lower() not in ["", "nan", "none"]:
                    status = "Report Received" if report_received else "Report Pending"
                else:
                    status = "Not Appointed"
                
                cw.writerow([
                    status, "Major", sheet_name, int(idx) + 1, cpf, name_designation, 
                    f"{dc_no_val} dt {dc_date_val}", eo_detail, 
                    "Received" if report_received else "Pending", report_detail,
                    "Issued" if scn_issued else "Pending", scn_detail
                ])
                
        output = si.getvalue()
        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-disposition": "attachment; filename=eo_reviews_report.csv"}
        )

    @app.get("/sync")
    @admin_required
    def sync_page():
        return render_page("sync.html")

    @app.post("/sync")
    @admin_required
    def run_sync():
        target_file = request.form.get("target_file", "").strip()
        if not target_file or not os.path.exists(target_file):
            return redirect(url_for("sync_page", message="Invalid file path.", status="error"))
            
        try:
            import pandas as pd
            updated_count = 0
            new_count = 0
            for sheet_name in ACTIVE_SEARCH_SHEETS:
                try:
                    circle_df = pd.read_excel(target_file, sheet_name=sheet_name, header=None, dtype=str)
                except Exception:
                    continue
                    
                circle_df = circle_df.fillna("")
                meta = loader.get_sheet_meta(sheet_name)
                start_row = max(meta.get("data_start", 5) - 1, 0)
                cpf_col = meta.get("cpf_col", 4) - 1
                dc_col = meta.get("dc_record_no_col", 11) - 1
                
                master_df = loader.load_dc_sheet(sheet_name, use_cache=False)
                
                for idx, row in circle_df.iloc[start_row:].iterrows():
                    row_values = list(row.values)
                    cpf = str(row_values[cpf_col]).strip() if cpf_col < len(row_values) else ""
                    dc = str(row_values[dc_col]).strip() if dc_col < len(row_values) else ""
                    
                    if not cpf and not dc:
                        continue
                        
                    match_found = False
                    master_idx = -1
                    if master_df is not None:
                        for m_idx, m_row in master_df.iloc[start_row:].iterrows():
                            m_cpf = str(m_row.iloc[cpf_col]).strip() if cpf_col < len(m_row) else ""
                            m_dc = str(m_row.iloc[dc_col]).strip() if dc_col < len(m_row) else ""
                            if (cpf and m_cpf == cpf) and (dc and m_dc == dc):
                                match_found = True
                                master_idx = m_idx
                                break
                                
                    record = {i+1: str(v) for i,v in enumerate(row_values)}
                    if match_found:
                        editor.update_record(sheet_name, master_idx + 1, record)
                        updated_count += 1
                    else:
                        editor.add_record(sheet_name, record)
                        new_count += 1
                        
            loader.clear_cache()
            return redirect(url_for("sync_page", message=f"Sync complete. Updated: {updated_count}, Added: {new_count}", status="success"))
        except Exception as e:
            return redirect(url_for("sync_page", message=f"Error syncing: {str(e)}", status="error"))

    @app.get("/monthly-editor")
    @login_required
    def monthly_editor():
        target_sheets_8dc = ["4DC", "7DC", "12DC", "14DC", "16DC", "20DC", "24DC", "34DC"]
        all_sheets = loader.op_sheets
        return render_page("monthly_editor.html", sheets_8dc=target_sheets_8dc, all_sheets=all_sheets)

    @app.get("/api/monthly-sheet-data")
    @login_required
    def api_monthly_sheet_data():
        sheet = request.args.get("sheet", "")
        if not sheet:
            return {"error": "Sheet parameter missing"}
        preview = editor.get_sheet_preview(sheet, max_rows=2000)
        if preview is None:
            return {"error": "Sheet not found"}
        
        headers = editor.get_column_labels(sheet)
        field_map = editor.get_sheet_field_map(sheet)
        return {"preview": {"headers": headers, "rows": preview}, "field_map": field_map}

    def validate_seniority(sheet_name, row_data, editor, loader):
        field_map = editor.get_sheet_field_map(sheet_name)
        desig_idx = None
        for key, col_idx in field_map.items():
            if "designation" in key.lower():
                desig_idx = col_idx
                break
                
        if desig_idx is not None and desig_idx in row_data:
            designation = str(row_data[desig_idx]).strip()
            
            if not designation:
                return f"Validation Error: Designation cannot be empty. Please wait for the auto-fetch or enter it manually."
                
            cpf_col = None
            for key, col_idx in field_map.items():
                if "cpfno" in key.lower() or "cpf" in key.lower():
                    cpf_col = col_idx
                    break
                    
            pg = ""
            if cpf_col is not None and cpf_col in row_data:
                emp = editor.get_employee_by_cpf(str(row_data[cpf_col]).strip())
                if emp:
                    pg = emp.get("paygrp", "")
            
            if not pg:
                pg_col = field_map.get("paygroup")
                if pg_col is not None and pg_col in row_data:
                    pg = str(row_data[pg_col]).strip()
                    
            seniority = loader.resolve_seniority(designation, pg)
            
            STATE_SHEETS = ["4DC", "7DC", "12DC", "14DC", "16DC", "20DC", "22DC", "24DC", "29DC"]
            is_state_sheet = sheet_name.strip().upper() in STATE_SHEETS
            is_state_emp = "state" in seniority.lower()
            
            if is_state_sheet and not is_state_emp:
                return f"Validation Error: Employee with designation '{designation}' has {seniority} seniority and MUST NOT be in a State Seniority sheet ({sheet_name})."
            if not is_state_sheet and is_state_emp:
                return f"Validation Error: Employee with designation '{designation}' has {seniority} seniority and MUST be in a State Seniority sheet (e.g. 4DC)."
        return None

    @app.post("/api/monthly-record/add")
    @login_required
    def api_monthly_record_add():
        data = request.json
        if not data or "sheet" not in data or "row_data" not in data:
            return {"error": "Invalid request payload"}
        try:
            parsed_data = {int(k): str(v) for k, v in data["row_data"].items()}
            err = validate_seniority(data["sheet"], parsed_data, editor, loader)
            if err:
                return {"error": err}
                
            res = editor.add_record(data["sheet"], parsed_data)
            loader.clear_sheet_cache(data["sheet"])
            
            # Auto-upload logic
            if res and "row_number" in res:
                if data["sheet"] in ("4DC", "5DC"):
                    editor.copy_record_to_sheet(data["sheet"], res["row_number"], "6DC")
                    loader.clear_sheet_cache("6DC")
                elif data["sheet"] == "20DC":
                    editor.copy_record_to_sheet(data["sheet"], res["row_number"], "22DC")
                    loader.clear_sheet_cache("22DC")
                elif data["sheet"] == "21DC":
                    editor.copy_record_to_sheet(data["sheet"], res["row_number"], "23DC")
                    loader.clear_sheet_cache("23DC")
                
            return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    @app.post("/api/monthly-record/update")
    @login_required
    def api_monthly_record_update():
        data = request.json
        if not data or "sheet" not in data or "row_number" not in data or "row_data" not in data:
            return {"error": "Invalid request payload"}
        try:
            parsed_data = {int(k): str(v) for k, v in data["row_data"].items()}
            err = validate_seniority(data["sheet"], parsed_data, editor, loader)
            if err:
                return {"error": err}
                
            editor.update_record(data["sheet"], int(data["row_number"]), parsed_data)
            loader.clear_sheet_cache(data["sheet"])
            return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    @app.get("/api/dispatch-prefixes")
    def api_dispatch_prefixes():
        try:
            import pandas as pd
            import os
            
            # Search for the right ofc file
            possible_paths = [
                os.path.join(os.path.dirname(loader.dc_file), 'ofc.xls.xlsx'),
                os.path.join(os.path.dirname(loader.dc_file), 'ofc.xlsx'),
                os.path.join(os.path.dirname(loader.dc_file), 'ofc.xls'),
                r'D:\MYPRO\HRMS\ofc.xls.xlsx',
                r'D:\MYPRO\HRMS\ofc.xlsx',
                r'D:\MYPRO\HRMS\ofc.xls'
            ]
            
            ofc_path = None
            for p in possible_paths:
                if os.path.exists(p):
                    ofc_path = p
                    break
                    
            if ofc_path:
                engine = 'openpyxl' if ofc_path.endswith('.xlsx') else 'xlrd'
                df = pd.read_excel(ofc_path, engine=engine)
                
                # Find locnm (column index 1 usually) and dispatch_sr
                loc_col = 'locnm' if 'locnm' in df.columns else df.columns[1] if len(df.columns) > 1 else None
                
                target_col = None
                for c in df.columns:
                    if 'dispatch_sr' in str(c).lower():
                        target_col = c
                        break
                        
                if not target_col and len(df.columns) >= 8:
                    target_col = df.columns[7]
                    
                prefixes = []
                if target_col and loc_col:
                    # Return list of dicts: label=locnm, value=dispatch_sr
                    for _, row in df.iterrows():
                        val = str(row[target_col]).strip()
                        label = str(row[loc_col]).strip()
                        if val and val.lower() != 'nan':
                            prefixes.append({"label": label, "value": val})
                    # Deduplicate by value
                    seen = set()
                    unique_prefixes = []
                    for p in prefixes:
                        if p['value'] not in seen:
                            seen.add(p['value'])
                            unique_prefixes.append(p)
                    return {"prefixes": unique_prefixes}
                elif target_col:
                    vals = df[target_col].dropna().astype(str).unique().tolist()
                    return {"prefixes": [{"label": v, "value": v} for v in vals]}
                else:
                    return {"prefixes": []}
            return {"prefixes": []}
        except Exception as e:
            return {"prefixes": [], "error": str(e)}

    @app.get("/monthly-export")
    @login_required
    def monthly_export():
        import openpyxl
        try:
            wb = openpyxl.load_workbook(loader.dc_file, data_only=False)
            target_sheets = {"4DC", "7DC", "12DC", "14DC", "16DC", "20DC", "24DC", "34DC", "INDEX", "Ltr"}
            keep_sheet_names = set()
            for ts in target_sheets:
                if ts in META and "actual_name" in META[ts]:
                    keep_sheet_names.add(META[ts]["actual_name"])
                else:
                    keep_sheet_names.add(ts)
            
            for sn in wb.sheetnames:
                if sn not in keep_sheet_names:
                    del wb[sn]
            
            month_year = datetime.now().strftime("%m%Y")
            file_name = f"PZ_8DC_{month_year}.xlsx"
            temp_path = os.path.join(loader.cache_dir, file_name)
            wb.save(temp_path)
            return send_file(temp_path, as_attachment=True, download_name=file_name)
        except Exception as e:
            return redirect(url_for("home", message=f"Export failed: {str(e)}", status="error"))

    @app.get("/reports/unmatched-employees")
    @login_required
    def unmatched_employees():
        from DC_DatabaseManager import db_manager
        try:
            with db_manager.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT d.sheet_origin, d.cpf_no, d.employee_name, d.designation, d.present_office, d.dc_number
                    FROM disciplinary_cases d
                    LEFT JOIN employees e ON d.cpf_no = e.cpf_no
                    WHERE e.cpf_no IS NULL AND d.cpf_no != '' AND d.cpf_no IS NOT NULL AND d.cpf_no != '0'
                    ORDER BY d.sheet_origin
                """)
                unmatched = [dict(r) for r in cursor.fetchall()]
        except Exception as e:
            return redirect(url_for("home", message=f"Database error: {str(e)}", status="error"))
            
        return render_page("unmatched.html", unmatched=unmatched)

    @app.get("/export_changes")
    @login_required
    def export_changes():
        try:
            from export_utils import generate_highlighted_export
            from DC_DataLoader import APP_DIR
            
            temp_path = generate_highlighted_export(loader.dc_file, APP_DIR)
            filename = os.path.basename(temp_path)
            
            return send_file(temp_path, as_attachment=True, download_name=filename)
        except Exception as e:
            return redirect(url_for("home", message=f"Export Changes failed: {str(e)}", status="error"))

    @app.get("/api/clear_audit_log")
    @login_required
    def api_clear_audit_log():
        try:
            from export_utils import clear_audit_log
            from DC_DataLoader import APP_DIR
            clear_audit_log(APP_DIR)
            return redirect(url_for("home", message="Audit Log cleared successfully!", status="success"))
        except Exception as e:
            return redirect(url_for("home", message=f"Failed to clear log: {str(e)}", status="error"))

    @app.get("/api/clear_cache")
    def api_clear_cache():
        try:
            loader.clear_cache()
            ref = request.referrer
            if ref:
                from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
                parsed = urlparse(ref)
                query_params = parse_qs(parsed.query)
                query_params['message'] = ['Data successfully updated!']
                query_params['status'] = ['success']
                new_query = urlencode(query_params, doseq=True)
                ref = urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))
                return redirect(ref)
            return redirect(url_for("home", message="Data successfully updated!", status="success"))
        except Exception as e:
            return redirect(url_for("home", message=f"Update failed: {str(e)}", status="error"))

    return app

def run_server(loader, host="0.0.0.0", port=5000, open_browser=True):
    app = create_app(loader)
    display_host = "127.0.0.1" if host == "0.0.0.0" else host
    url = f"http://{display_host}:{port}/login"
    if open_browser:
        try:
            webbrowser.open(url)
        except Exception:
            pass
    app.run(host=host, port=port, debug=False)
